"""
NLSC 臺灣通用電子地圖 異動情資蒐整 Demo 程式
================================================
動態版：所有關鍵字與新聞均即時衍生，無任何硬寫資料。

資料來源：
  - 關鍵字.txt     → 工程納入 / 排除詞彙
  - NLSC 管控表單  → 既有案件位置詞彙 + 已管控案件名單
  - 每日新聞 JSON  → 新聞來源 A（每日更新）
  - Google News RSS → 新聞來源 B（即時 Web Search）
  - 工程會 JSON    → 工程會清冊處理（測試 1~4）

執行方式：
    python demo_emap_intelligence.py
"""

import json
import re
import sys
from difflib import SequenceMatcher
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# 0. 設定區（唯一需要調整的地方）
# ─────────────────────────────────────────────────────────────
BASE_DIR        = Path(__file__).parent
JSON_127        = BASE_DIR / "output_file_all_case_format_v2_20260127.json"
JSON_415        = BASE_DIR / "output_file_all_case_format_v2_20260415.json"
NLSC_FORM       = BASE_DIR / "115EMAP案管_1150422未結案案件.xlsx"
PCC_CASE_MGR    = BASE_DIR / "案件管理(上半年篩選)_11503版_提交甲方_0324_FIN_nlsc.xlsx"
KEYWORD_FILE    = BASE_DIR / "異動資料蒐集37月更新_關鍵字_V1140320.txt"
NEWS_JSON_GLOB  = "20260503_vw_8DNewsAI_noTag新聞資料.json"   # 每日 JSON 檔名 glob

OUT_PCC   = BASE_DIR / "demo_工程會處理結果_北北桃.xlsx"
OUT_NEWS  = BASE_DIR / "demo_新聞爬搜結果_北北桃.xlsx"

# ── 統一搜尋範圍：工程會清冊 + 新聞爬搜 共用同一設定 ──
NEWS_CITIES      = ["臺北市", "桃園市"]   # 涵蓋縣市（工程會+新聞同步）
NEWS_CITIES_SHORT= ["臺北", "台北", "桃園"]  # 縮寫（供位置關鍵字比對）

# 向下相容（部分顯示文字用）
TARGET_CITY     = NEWS_CITIES[0]
TARGET_DISTRICT = ""

MIN_PRICE_K      = 5_000           # 500 萬元（千元單位）
NEWS_DAYS        = 30              # 近幾天 = 新聞 JSON 過濾窗口
WEB_SEARCH_QTY   = 10             # Web Search 每次查詢最多回傳筆數

# ── LLM 後分類設定（自家 API，可選）───────────────────────
_LLM_CONFIG_FILE = BASE_DIR / ".emap_llm_config.json"
def _load_llm_cfg() -> dict:
    cfg = {}
    if _LLM_CONFIG_FILE.exists():
        import json as _j
        with open(_LLM_CONFIG_FILE, encoding="utf-8") as _f:
            cfg = _j.load(_f)
    import os
    return {
        "api_key":  str(cfg.get("api_key")  or os.getenv("EMAP_LLM_API_KEY")  or "").strip(),
        "base_url": str(cfg.get("base_url") or os.getenv("EMAP_LLM_BASE_URL")
                        or "http://125.227.53.125:50062/").strip(),
        "model":    str(cfg.get("model")    or os.getenv("EMAP_LLM_MODEL")    or "gpt-oss-120b").strip(),
        "timeout":  int(cfg.get("timeout")  or 30),
    }
LLM_CFG = _load_llm_cfg()
PROXIMITY_CHARS  = 300             # 位置詞 & 工程詞需在此字元範圍內共現
WEB_DEBUG        = True            # 顯示 Web Search 過濾細節
LLM_DEDUP        = True            # 新聞去重優先使用 LLM 判斷是否同一工程事件
_LLM_DEDUP_CACHE: dict[str, tuple[bool, str]] = {}


# ─────────────────────────────────────────────────────────────
# 1. 動態載入關鍵字
# ─────────────────────────────────────────────────────────────
def load_keywords(filepath: Path):
    text = filepath.read_text(encoding="utf-8")
    def _extract(section):
        m = re.search(rf'{section}\s*=\s*\[(.*?)\]', text, re.DOTALL)
        return re.findall(r'"([^"]+)"', m.group(1)) if m else []

    return {
        "include":       _extract("應納入關鍵字"),
        "exclude":       _extract("非納入關鍵字"),
        "include_units": _extract("應納入單位"),
        "exclude_units": _extract("非納入單位"),
    }


KW = load_keywords(KEYWORD_FILE)


# ── 關鍵字 → regex 轉換（glob-like 寫法安全化）─────────────────
def _kw_to_regex(kw: str) -> str:
    """
    關鍵字檔用 glob-like 寫法，直接給 re.search 會誤判：
      "台*線" → 台[0-9]+線  （台1線、台9線、台61線…）
      "K+"    → K\+[0-9]   （K+里程標，如 K+500）
      其他    → re.escape   （字面比對）
    """
    if kw == "台*線":
        return r"台[0-9]+線"
    if kw == "K+":
        return r"K\+[0-9]"
    return re.escape(kw)

_KW_REGEX_CACHE: dict[str, re.Pattern] = {}

def _kw_search(kw: str, text: str) -> bool:
    """以安全 regex 搜尋關鍵字是否出現在 text 中"""
    if kw not in _KW_REGEX_CACHE:
        _KW_REGEX_CACHE[kw] = re.compile(_kw_to_regex(kw))
    return bool(_KW_REGEX_CACHE[kw].search(text))


# 工程篩選用函式（工程會清冊）
def kw_pass(name: str) -> bool:
    return (any(_kw_search(k, name) for k in KW["include"])
            and not any(k in name for k in KW["exclude"]))

def unit_pass(host: str) -> bool:
    if not host:
        return True
    return not any(k in host for k in KW["exclude_units"])


# ─────────────────────────────────────────────────────────────
# 2. 動態衍生位置關鍵字（從設定 + NLSC 管控表單）
# ─────────────────────────────────────────────────────────────
_ROAD_RE = re.compile(r'[一-鿿]{2,6}(?:路|街|巷|段|大道|橋|路口)')

def derive_loc_keywords(city: str, district: str, nlsc_df: pd.DataFrame) -> list[str]:
    """
    工程會清冊用：衍生單一縣市+區的位置關鍵字（含 NLSC 門牌路名）
    """
    base = [district, f"{district}區", f"{city}{district}", f"{city}{district}區",
            "內科"]   # 「內科」= 內湖科學園區，通用縮稱保留

    road_words = set()
    taipei_rows = nlsc_df[nlsc_df.get("坐落_縣市", pd.Series()).astype(str)
                          .str.contains(city, na=False)]
    for val in taipei_rows.get("核對追蹤_參考門牌", pd.Series()).dropna():
        for m in _ROAD_RE.findall(str(val)):
            road_words.add(m)

    combined = base + sorted(road_words, key=len)
    return combined


def derive_news_loc_keywords(cities: list[str], cities_short: list[str],
                             nlsc_df: pd.DataFrame) -> list[str]:
    """
    新聞爬搜用：衍生多縣市位置關鍵字（不限區域），包含縮寫與 NLSC 路名
    """
    base = list(cities) + list(cities_short)   # 臺北市, 桃園市, 臺北, 台北, 桃園

    road_words = set()
    for city in cities:
        rows = nlsc_df[nlsc_df.get("坐落_縣市", pd.Series()).astype(str)
                       .str.contains(city[:2], na=False)]
        for val in rows.get("核對追蹤_參考門牌", pd.Series()).dropna():
            for m in _ROAD_RE.findall(str(val)):
                road_words.add(m)

    return base + sorted(road_words, key=len)


# ─────────────────────────────────────────────────────────────
# 3. 工具函式
# ─────────────────────────────────────────────────────────────
def roc_to_ad(s: str) -> str:
    if not s:
        return ""
    m = re.match(r'^(\d{3})(\d{2})(\d{2})$', str(s).strip())
    return f"{int(m.group(1))+1911}-{m.group(2)}-{m.group(3)}" if m else str(s).strip()


def _extract_progress(text: str) -> str:
    for kw, label in [
        ("啟用", "完工啟用"), ("揭牌", "完工揭牌"), ("落成", "完工落成"),
        ("竣工", "竣工"), ("通車", "通車"), ("完工", "施工完成"),
        ("動土", "開工（動土典禮）"), ("動工", "開工"), ("開工", "開工"),
        ("改建中", "施工中（改建）"), ("施工", "施工中"),
        ("招標", "規劃設計（招標中）"), ("規劃", "規劃設計"),
    ]:
        if kw in text:
            return label
    return "施工中"


def _extract_completion(text: str) -> str:
    for pat in [
        r'(\d{4})[年/\-](\d{1,2})[月/\-](\d{1,2})[日]?',
        r'(\d{4})年(\d{1,2})月',
        r'民國(\d{3})年(\d{1,2})月(\d{1,2})日',
        r'預計(\d{4})年.*?完工',
    ]:
        m = re.search(pat, text)
        if m:
            g = m.groups()
            y = int(g[0]) + (1911 if int(g[0]) < 200 else 0)
            if len(g) >= 3 and g[2]:
                return f"{y}-{int(g[1]):02d}-{int(g[2]):02d}"
            elif len(g) >= 2:
                return f"{y}-{int(g[1]):02d}"
    return ""


def _extract_case_name(title: str, content: str) -> str:
    """
    從標題 / 內文萃取案名，優先「XXX工程」型，
    次之為常見設施名稱模式，最後取標題前段。
    """
    # 1. 「XXX工程」型
    m = re.search(r'[一-鿿（）()A-Za-z0-9\s]{4,35}工程', title)
    if m:
        return m.group(0).strip()

    # 2. 設施名稱模式（動態，不硬寫清單）
    full = title + " " + content[:300]
    FACILITY_PATS = [
        (r'([一-鿿]{2,10}(?:公園|廣場|體育館|圖書館|社宅|社會住宅))', "新建工程"),
        (r'([一-鿿]{2,8}(?:大橋|陸橋|景觀橋|跨河橋))', "橋梁工程"),
        (r'(捷運[一-鿿 A-Za-z0-9]{2,15}(?:站|段|標))', "捷運工程"),
        (r'([一-鿿]{2,10}(?:變電所|發電廠|機房))', "新建工程"),
        (r'([一-鿿]{2,10}(?:大樓|廳舍|校舍|宿舍))', "新建工程"),
    ]
    for pat, suffix in FACILITY_PATS:
        m = re.search(pat, full)
        if m:
            facility = m.group(1).strip()
            if len(facility) >= 3:
                return f"{facility}{suffix}"

    # 3. 標題前段（去掉記者署名、媒體前綴等）
    cleaned = re.sub(r'^[\[【〔「].*?[\]】〕」]\s*', '', title)
    return cleaned[:40].strip()


def _extract_case_name_candidates(title: str, content: str, max_cases: int = 8) -> list[str]:
    """
    單篇新聞可拆多案件：先抓「XXX工程」與路線型案件（台61線/台62線），
    若仍抓不到再退回既有單案名邏輯。
    """
    full = f"{title} {content[:3500]}"
    candidates: list[str] = []

    # 1) 直接抓「XXX工程」
    for m in re.finditer(r'[一-鿿（）()A-Za-z0-9、，,\-／/\s]{4,40}工程', full):
        name = re.sub(r'\s+', '', m.group(0)).strip("，。、；：")
        if 4 <= len(name) <= 26:
            candidates.append(name)

    # 2) 路線型案件（如 台61線、台62線）可拆為多筆
    route_hits = list(dict.fromkeys(re.findall(r'台[0-9]{1,3}線', full)))
    action_words = ["擴建", "延伸", "新建", "改善", "拓寬", "改建", "高架化", "南延", "西延", "東延"]
    for route in route_hits:
        idx = full.find(route)
        window = full[max(0, idx - 30): idx + 80] if idx >= 0 else full[:120]
        action = next((w for w in action_words if w in window), "")
        suffix = f"{action}工程" if action else "工程"
        candidates.append(f"{route}{suffix}")

    if not candidates:
        single = _extract_case_name(title, content)
        return [single] if single else []

    unique: list[str] = []
    seen = set()
    for c in candidates:
        key = _dedup_key(c) or c
        if key in seen:
            continue
        seen.add(key)
        unique.append(c)
        if len(unique) >= max_cases:
            break
    return unique


def _extract_news_article_from_url(url: str) -> dict:
    """抓取單篇新聞網址並萃取標題、日期、內文。"""
    import html as _html
    import urllib.parse as _up
    import urllib.request as _ur

    req = _ur.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept-Language": "zh-TW,zh;q=0.9",
        },
    )
    with _ur.urlopen(req, timeout=15) as resp:
        page = resp.read().decode("utf-8", errors="ignore")

    def _meta_content(*keys: str) -> str:
        for key in keys:
            pat1 = rf'<meta[^>]+(?:property|name)=["\']{re.escape(key)}["\'][^>]*content=["\']([^"\']+)["\']'
            pat2 = rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:property|name)=["\']{re.escape(key)}["\']'
            m = re.search(pat1, page, flags=re.IGNORECASE)
            if not m:
                m = re.search(pat2, page, flags=re.IGNORECASE)
            if m:
                return _html.unescape(m.group(1)).strip()
        return ""

    title = _meta_content("og:title", "twitter:title", "title")
    if not title:
        m = re.search(r'<title>(.*?)</title>', page, flags=re.IGNORECASE | re.DOTALL)
        title = _html.unescape(m.group(1)).strip() if m else ""
    title = _clean_news_title(title)

    published_raw = _meta_content("article:published_time", "pubdate", "datePublished")
    dm = re.search(r'(\d{4}-\d{2}-\d{2})', published_raw or "")
    date_str = dm.group(1) if dm else str(date.today())

    # 優先抓文章主體 editor 區塊，避免吃到延伸閱讀
    block = ""
    m_editor = re.search(
        r'<section[^>]*class="[^"]*article-content__editor[^"]*"[^>]*>(.*?)</section>',
        page, flags=re.IGNORECASE | re.DOTALL
    )
    if m_editor:
        block = m_editor.group(1)
    else:
        m_article = re.search(
            r'<article[^>]*class="[^"]*article-content[^"]*"[^>]*>(.*?)</article>',
            page, flags=re.IGNORECASE | re.DOTALL
        )
        if m_article:
            block = m_article.group(1)

    paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', block, flags=re.IGNORECASE | re.DOTALL) if block else []
    body_parts: list[str] = []
    for p in paragraphs:
        txt = re.sub(r'<[^>]+>', '', p)
        txt = _html.unescape(txt).strip()
        txt = re.sub(r'\s+', ' ', txt)
        if len(txt) >= 8:
            body_parts.append(txt)
    body = " ".join(body_parts)
    if len(body) < 30:
        body = _meta_content("og:description", "description")

    source = _up.urlparse(url).netloc or "手動網址"
    return {
        "title": title,
        "body": body,
        "url": url,
        "source": source,
        "date": date_str,
    }


def _loc_build_proximity(text: str, loc_kws: list, build_kws: list) -> bool:
    """位置關鍵字與工程關鍵字需在 PROXIMITY_CHARS 字元內共現"""
    for loc in loc_kws:
        idx = text.find(loc)
        if idx < 0:
            continue
        window = text[max(0, idx - PROXIMITY_CHARS): idx + PROXIMITY_CHARS]
        if any(_kw_search(bk, window) for bk in build_kws):
            return True
    return False


def _news_has_build(title: str, body: str, build_kws: list) -> bool:
    """
    新聞專用工程詞比對（比工程會清冊更嚴格）：

    應納入關鍵字設計上是篩「工程案名」短字串，直接套到新聞全文太寬。
    本函式把關鍵字分兩級：

      ‣ 複合詞（≥4字）：如「新建工程」「道路拓寬工程」「市地重劃」
        → 出現在標題或內文皆算命中（夠具體，不易誤判）

      ‣ 短詞（<4字）：如「新建」「興建」「大橋」「國道」
        → 必須出現在標題，或後方緊接「工程／計畫／道路」才算
        → 避免「無線電遭盜接」「板南線房市」等誤判
    """
    for kw in build_kws:
        if len(kw) >= 4:
            if _kw_search(kw, title + body):
                return True
        else:
            # 短詞：標題出現，或內文中後接工程相關詞
            if _kw_search(kw, title):
                return True
            suffix_pat = _kw_to_regex(kw) + r'(?:工程|計畫|道路|路線|設施|橋梁|改建|新建)'
            if re.search(suffix_pat, body):
                return True
    return False


def _clean_url(url: str) -> str:
    raw = str(url or "").strip()
    if not raw:
        return ""
    try:
        import urllib.parse as _up
        sp = _up.urlsplit(raw)
        query_pairs = _up.parse_qsl(sp.query, keep_blank_values=True)
        drop_keys = {
            "utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term",
            "fbclid", "gclid", "igshid", "mkt_tok",
        }
        kept = [(k, v) for k, v in query_pairs if k.lower() not in drop_keys]
        query = _up.urlencode(kept, doseq=True)
        path = sp.path.rstrip("/")
        return _up.urlunsplit((sp.scheme.lower(), sp.netloc.lower(), path, query, ""))
    except Exception:
        return raw


def _extract_llm_content(resp_obj: dict) -> str:
    """讀取 OpenAI-compatible chat completion 的文字回覆；content 為 null 時回傳空字串。"""
    msg = (resp_obj.get("choices") or [{}])[0].get("message") or {}
    content = msg.get("content")
    return content.strip() if isinstance(content, str) else ""


def _llm_is_construction(title: str, body: str) -> tuple[bool, str]:
    """
    用自家 LLM 判斷新聞是否為工程建設類（通用後分類）。
    回傳 (is_construction, reason_str)
    若未設定 API key 或呼叫失敗，預設放行（True）。
    """
    if not LLM_CFG["api_key"]:
        return True, "（未設定 LLM key，跳過）"

    import urllib.request as _ur
    prompt = (
        "不要解釋，不要推理。\n"
        "判斷以下新聞是否描述「道路、橋梁或建築物的新建、改建、拓寬、完工、開工」等工程建設事件。\n"
        f"標題：{title}\n內文摘要：{body[:300]}\n\n"
        "只回答：是｜理由  或  否｜理由（一句話）"
    )
    payload = json.dumps({
        "model": LLM_CFG["model"], "temperature": 0, "max_tokens": 300,
        "messages": [{"role": "user", "content": prompt}]
    }, ensure_ascii=False).encode("utf-8")

    base = LLM_CFG["base_url"].rstrip("/")
    if not base.endswith("/v1"):
        base += "/v1"
    req = __import__("urllib.request", fromlist=["Request"]).Request(
        f"{base}/chat/completions", data=payload,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {LLM_CFG['api_key']}"},
        method="POST",
    )
    try:
        with _ur.urlopen(req, timeout=LLM_CFG["timeout"]) as resp:
            content = _extract_llm_content(json.loads(resp.read()))
        if not content:
            return True, "（LLM空回覆，預設放行）"
        return content.startswith("是"), content
    except Exception as e:
        return True, f"（LLM失敗:{type(e).__name__}，預設放行）"


def _dedup_key(case_name: str) -> str:
    """文字備援用桶鍵；LLM 不可用時才作為主要去重依據。"""
    # 保留數字（例如台61線/台62線）避免不同路線被誤合併
    key = re.sub(r'[A-Za-z（）()！？!?:：.,，。、\s]', '', case_name)
    return key[:20]


def _route_tags(text: str) -> set[str]:
    return set(re.findall(r'台[0-9]{1,3}線', str(text or "")))


def _clean_news_title(title: str) -> str:
    """去掉 RSS 標題常見的媒體尾巴，讓 LLM 聚焦在工程事件本身。"""
    text = str(title or "").strip()
    text = re.sub(r'\s*[-－–]\s*[^-－–]{2,20}$', '', text)
    text = re.sub(r'\s*[|｜]\s*[^|｜]{2,20}$', '', text)
    return text.strip()


def _news_brief(item: dict, body_key: str) -> str:
    body = str(item.get(body_key, "") or "")
    parts = [
        f"標題：{_clean_news_title(item.get('_title', ''))}",
        f"原標題：{item.get('_title', '')}",
        f"案名：{item.get('_case', '')}",
        f"日期：{item.get('_date', '')}",
        f"來源：{item.get('_source', '')}",
        f"摘要：{body[:420]}",
    ]
    return "\n".join(parts)


def _fallback_same_news_event(a: dict, b: dict) -> tuple[bool, str]:
    """LLM 不可用時的保守備援：案名/標題核心高度重疊才合併。"""
    a_case, b_case = a.get("_case", ""), b.get("_case", "")
    a_title, b_title = _clean_news_title(a.get("_title", "")), _clean_news_title(b.get("_title", ""))
    ra_case, rb_case = _route_tags(a_case), _route_tags(b_case)
    if ra_case and rb_case and ra_case.isdisjoint(rb_case):
        return False, f"fallback: 案名路線不同({sorted(ra_case)} vs {sorted(rb_case)})"
    ra = ra_case or _route_tags(a_title)
    rb = rb_case or _route_tags(b_title)
    if ra and rb and ra.isdisjoint(rb):
        return False, f"fallback: 路線不同({sorted(ra)} vs {sorted(rb)})"
    if _dedup_key(a_case) and _dedup_key(a_case) == _dedup_key(b_case):
        return True, "fallback: 案名桶鍵相同"
    score = max(_overlap_score(a_case, b_case), _overlap_score(a_title, b_title))
    return score >= 0.92, f"fallback: text_score={score:.2f}"


def _high_confidence_same_news_event(a: dict, b: dict) -> tuple[bool, str]:
    """LLM 判不同時的高信心保護：同日且核心標題/案名高度重疊才合併。"""
    a_case, b_case = a.get("_case", ""), b.get("_case", "")
    a_title, b_title = _clean_news_title(a.get("_title", "")), _clean_news_title(b.get("_title", ""))
    ra_case, rb_case = _route_tags(a_case), _route_tags(b_case)
    if ra_case and rb_case and ra_case.isdisjoint(rb_case):
        return False, f"案名路線不同({sorted(ra_case)} vs {sorted(rb_case)})"
    ra = ra_case or _route_tags(a_title)
    rb = rb_case or _route_tags(b_title)
    if ra and rb and ra.isdisjoint(rb):
        return False, f"路線不同({sorted(ra)} vs {sorted(rb)})"
    score = max(_overlap_score(a_case, b_case), _overlap_score(a_title, b_title))
    same_date = str(a.get("_date", "") or "")[:10] == str(b.get("_date", "") or "")[:10]
    if score >= 0.9 or (same_date and score >= 0.75):
        return True, f"LLM後高信心合併: text_score={score:.2f}, same_date={same_date}"
    return False, f"text_score={score:.2f}, same_date={same_date}"


def _llm_same_news_event(a: dict, b: dict, body_key: str) -> tuple[bool, str]:
    """
    用 LLM 判斷兩則新聞是否為同一工程案件/同一異動事件。
    這裡刻意提供標題、案名、日期、來源、摘要，不只看標題。
    """
    a_url = _clean_url(a.get("_url", ""))
    b_url = _clean_url(b.get("_url", ""))
    if a_url and b_url and a_url == b_url:
        a_case, b_case = str(a.get("_case", "") or ""), str(b.get("_case", "") or "")
        ra_case, rb_case = _route_tags(a_case), _route_tags(b_case)
        if ra_case and rb_case and ra_case.isdisjoint(rb_case):
            return False, f"同網址但案名路線不同({sorted(ra_case)} vs {sorted(rb_case)})"
        case_score = _overlap_score(a_case, b_case)
        if (_dedup_key(a_case) and _dedup_key(a_case) == _dedup_key(b_case)) or case_score >= 0.86:
            return True, f"同網址且案名相近(case_score={case_score:.2f})"
        return False, f"同網址但案名不同(case_score={case_score:.2f})"

    if not LLM_DEDUP or not LLM_CFG["api_key"]:
        return _fallback_same_news_event(a, b)

    cache_key = "||".join(sorted([
        str(a.get("_url") or a.get("_title") or ""),
        str(b.get("_url") or b.get("_title") or ""),
    ]))
    if cache_key in _LLM_DEDUP_CACHE:
        return _LLM_DEDUP_CACHE[cache_key]

    import urllib.request as _ur
    prompt = (
        "不要解釋，不要推理。\n"
        "請判斷以下兩則新聞是否描述同一個工程案件或同一個工程進度異動事件。\n"
        "判斷時要看案名、地點、工程內容、日期與摘要，不可只看標題。\n"
        "若只是同縣市、同類工程、同道路類型，不能算同一案。\n\n"
        "新聞A：\n"
        f"{_news_brief(a, body_key)}\n\n"
        "新聞B：\n"
        f"{_news_brief(b, body_key)}\n\n"
        "只回答：同一｜理由  或  不同｜理由（一句話）"
    )
    payload = json.dumps({
        "model": LLM_CFG["model"], "temperature": 0, "max_tokens": 360,
        "messages": [{"role": "user", "content": prompt}]
    }, ensure_ascii=False).encode("utf-8")

    base = LLM_CFG["base_url"].rstrip("/")
    if not base.endswith("/v1"):
        base += "/v1"
    req = __import__("urllib.request", fromlist=["Request"]).Request(
        f"{base}/chat/completions", data=payload,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {LLM_CFG['api_key']}"},
        method="POST",
    )
    try:
        with _ur.urlopen(req, timeout=LLM_CFG["timeout"]) as resp:
            content = _extract_llm_content(json.loads(resp.read()))
        if not content:
            same, reason = _high_confidence_same_news_event(a, b)
            if not same:
                same, reason = _fallback_same_news_event(a, b)
            result = same, f"LLM空回覆; {reason}"
            _LLM_DEDUP_CACHE[cache_key] = result
            return result
        if content.startswith("同一"):
            result = True, content
        else:
            same, reason = _high_confidence_same_news_event(a, b)
            result = (True, reason) if same else (False, content)
    except Exception as e:
        same, reason = _high_confidence_same_news_event(a, b)
        if not same:
            same, reason = _fallback_same_news_event(a, b)
        result = same, f"LLM去重失敗:{type(e).__name__}; {reason}"

    _LLM_DEDUP_CACHE[cache_key] = result
    return result


def _prefer_news_candidate(new_item: dict, old_item: dict, body_key: str) -> bool:
    """同一事件保留較新日期；日期相同時保留摘要較完整者。"""
    new_date = str(new_item.get("_date", "") or "")
    old_date = str(old_item.get("_date", "") or "")
    if new_date != old_date:
        return new_date > old_date
    return len(str(new_item.get(body_key, "") or "")) > len(str(old_item.get(body_key, "") or ""))


def _resolve_news_bucket(candidate: dict, buckets: dict[str, dict], body_key: str) -> tuple[str, str]:
    """
    針對已保留候選逐筆做 LLM 去重。
    回傳 bucket_key 與判斷理由；若文字桶鍵碰撞但 LLM 判定不同，會建立新桶。
    """
    for key, existing in buckets.items():
        same, reason = _llm_same_news_event(candidate, existing, body_key)
        if same:
            return key, reason

    base_key = _dedup_key(candidate.get("_case", "")) or _dedup_key(_clean_news_title(candidate.get("_title", ""))) or "新聞"
    key = base_key
    counter = 2
    while key in buckets:
        key = f"{base_key}#{counter}"
        counter += 1
    mode = "LLM新桶" if LLM_DEDUP and LLM_CFG["api_key"] else "fallback新桶"
    return key, mode


# ─────────────────────────────────────────────────────────────
# 4. 工程會 JSON 處理
# ─────────────────────────────────────────────────────────────
def load_prev_screened_names(filepath: Path) -> set[str]:
    """
    2-1 廠商初篩邏輯——從案件管理 Excel 讀取「前次(114下半年)篩選Y者」
    及「已確認納入控管者」的案名集合。

    判斷依據（欄位索引，從 0 起算）：
      col[2]  廠商初篩（115上半年）= 'Y'
      col[4]  114下半年篩選成果    = 'Y'
      col[5]  已納入控管           = 有值（非空）
      col[11] 基本資料_案名

    符合其中任一者 → 已處理，不須新增，更新狀態即可。
    """
    if not filepath.exists():
        print(f"⚠ 找不到案件管理 Excel：{filepath.name}，跳過前次篩選比對")
        return set()

    import openpyxl
    wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws   = wb.active
    names: set[str] = set()

    for row in ws.iter_rows(min_row=4, values_only=True):   # 前3列為標題
        c3  = str(row[2] or "").strip()   # 廠商初篩（115上半年）
        c5  = str(row[4] or "").strip()   # 114下半年篩選成果
        c6  = str(row[5] or "").strip()   # 已納入控管
        c12 = str(row[11] or "").strip()  # 基本資料_案名
        if not c12:
            continue
        if c3 == "Y" or c5 == "Y" or c6:
            names.add(c12)

    print(f"  → 前次篩選Y / 已納入控管案件：{len(names)} 筆（from {filepath.name}）")
    return names


def load_pcc_json(filepath: Path) -> list:
    with open(filepath, encoding="utf-8") as f:
        return json.load(f)


def filter_target(data: list) -> list:
    """篩選符合 NEWS_CITIES 範圍 + 金額 + 關鍵字的工程會案件"""
    out = []
    for item in data:
        bi       = item.get("basicInformation", {})
        district = bi.get("district", "")
        # 縣市比對：district 欄位通常如 "台北市內湖區"，取前2字比對
        if not any(c[:2] in district for c in NEWS_CITIES):
            continue
        price = bi.get("awardingPrice") or bi.get("totalBudget") or 0
        if price < MIN_PRICE_K:
            continue
        if not kw_pass(item.get("projectName", "")):
            continue
        if not unit_pass(bi.get("hostEntity", "")):
            continue
        out.append(item)
    return out


def extract_case(item: dict, report_date: str) -> dict:
    bi   = item.get("basicInformation", {})
    prog = item.get("progressions", [])
    latest = prog[-1] if prog else {}
    cp   = bi.get("contactPerson", {})
    return {
        "工程會案號":             item.get("projectNo", ""),
        "基本資料_案名":          item.get("projectName", ""),
        "坐落_縣市":              next((c for c in NEWS_CITIES if c[:2] in bi.get("district","")), NEWS_CITIES[0]),
        "坐落_區":                bi.get("district", ""),
        "基本資料_通報日期":      report_date,
        "基本資料_通報來源":      "服務雲",
        "基本資料_聯絡電話":      cp.get("telephoneNo", ""),
        "基本資料_電子郵件":      cp.get("emailAddress", ""),
        "核對追蹤_負責人":        cp.get("name", ""),
        "核對追蹤_執行單位":      bi.get("hostEntity", ""),
        "核對追蹤_參考門牌":      bi.get("location", ""),
        "核對追蹤_預算金額【千元】": bi.get("awardingPrice") or bi.get("totalBudget") or 0,
        "核對追蹤_進度":          latest.get("actualOverallProgress", ""),
        "核對追蹤_(預計)完工日":  roc_to_ad(bi.get("scheduledCompletionDate", "")),
        "核對追蹤_工程現況":      latest.get("summary", ""),
        "核對追蹤_案件備註":      bi.get("constructionSummary", ""),
        "案件摘要":               "",
        "篩選結果":               "",
        "篩選說明":               "",
    }


def auto_summary(case: dict) -> str:
    name     = case["基本資料_案名"]
    host     = case["核對追蹤_執行單位"]
    budget   = case["核對追蹤_預算金額【千元】"]
    location = case["核對追蹤_參考門牌"]
    progress = case["核對追蹤_進度"]
    compl    = case["核對追蹤_(預計)完工日"]
    notes    = case["核對追蹤_案件備註"] or case["核對追蹤_工程現況"]

    budget_s   = f"預算約 {budget/1000:.0f} 百萬元" if budget else ""
    progress_s = (f"，目前進度 {progress:.1f}%"
                  if isinstance(progress, (int, float)) else "")
    compl_s    = f"，預計完工 {compl}" if compl else ""
    loc_s      = f"，位於 {location}" if location else ""
    note_s     = f"。工程概要：{str(notes)[:80]}…" if notes else ""
    return f"{name}，由{host}執行{loc_s}。{budget_s}{progress_s}{compl_s}{note_s}".strip()


# ─────────────────────────────────────────────────────────────
# 5. 新聞 JSON 解析（來源 A）
# ─────────────────────────────────────────────────────────────
def parse_news_json(loc_kws: list, build_kws: list, noise_kws: list) -> tuple[list, list]:
    """
    掃描 BASE_DIR 下所有符合 NEWS_JSON_GLOB 的每日新聞 JSON，
    以動態關鍵字過濾並回傳結果。
    回傳 (results, dedup_track)：results=去重後清單，dedup_track=去重過程追蹤（供 Sheet 4）
    """
    files = sorted(BASE_DIR.glob(NEWS_JSON_GLOB))
    if not files:
        print(f"找不到新聞 JSON（{NEWS_JSON_GLOB}）")
        return [], []

    cutoff = date.today() - timedelta(days=NEWS_DAYS)
    # 桶鍵 → 內容最豐富的那篇
    buckets: dict[str, dict] = {}
    dedup_track: list[dict] = []   # 去重過程追蹤（LLM 判斷 + 去重結果）
    track_seq = 0

    for fp in files:
        print(f"讀取新聞 JSON：{fp.name}")
        with open(fp, encoding="utf-8") as f:
            raw = json.load(f)

        for item in raw:
            title    = item.get("DOCUMENT_TITLE", "")
            content  = item.get("DOCUMENT_CONTENT", "")
            url      = item.get("DOCUMENT_URL", "")
            source   = item.get("EXTRACTOR_NAME", "")
            date_str = item.get("DOCUMENT_DATE", "")[:10]

            try:
                doc_date = date.fromisoformat(date_str)
            except ValueError:
                continue
            if doc_date < cutoff:
                continue

            text = title + content
            if any(nk in text for nk in noise_kws):
                continue
            has_loc   = any(lk in text for lk in loc_kws)
            has_build = _news_has_build(title, content, build_kws)
            if not (has_loc and has_build):
                continue
            if not _loc_build_proximity(text, loc_kws, build_kws):
                continue

            # ── LLM 後分類（通用去噪：排除交通事故、政治新聞等）──
            ok, reason = _llm_is_construction(title, content)
            case_name = _extract_case_name(title, content)

            if not ok:
                track_seq += 1
                if WEB_DEBUG:
                    print(f"    [LLM-no] {title[:40]} → {reason[:40]}")
                dedup_track.append({
                    "track_id": track_seq, "title": title, "date": date_str, "content_len": len(content),
                    "llm_label": "❌ 非工程", "bucket_key": "—", "dedup_reason": reason, "result": "LLM排除",
                })
                continue

            track_seq += 1
            candidate = {
                "_content": content, "_title": title, "_url": url,
                "_source": source, "_date": date_str, "_case": case_name,
                "_track_id": track_seq,
            }
            bkt_key, dedup_reason = _resolve_news_bucket(candidate, buckets, "_content")

            dedup_track.append({
                "track_id": track_seq, "title": title, "date": date_str, "content_len": len(content),
                "llm_label": "✅ 是工程新聞", "bucket_key": bkt_key,
                "dedup_reason": dedup_reason, "result": "PENDING",
            })

            # 同一事件保留最新；同日保留內容較完整者
            existing = buckets.get(bkt_key)
            if existing is None or _prefer_news_candidate(candidate, existing, "_content"):
                buckets[bkt_key] = candidate

    # 確定 dedup_track 各筆最終結果（桶的唯一勝出者 = _title 在 buckets 中）
    winner_ids = {v.get("_track_id") for v in buckets.values()}
    for entry in dedup_track:
        if entry["result"] == "PENDING":
            entry["result"] = "✅ 保留" if entry.get("track_id") in winner_ids else "🔀 去重"

    results = []
    for bkt in buckets.values():
        content   = bkt["_content"]
        case_name = bkt["_case"]
        matched   = next((lk for lk in loc_kws if lk in (bkt["_title"] + content)), NEWS_CITIES[0])
        idx       = content.find(matched)
        loc_snip  = content[max(0,idx-5): idx+60].strip().replace("\n"," ") if idx >= 0 else matched
        city_hit  = next((c for c in NEWS_CITIES if c in bkt["_title"] + content), NEWS_CITIES[0])

        results.append({
            "新聞標題":     bkt["_title"],
            "案名":         case_name,
            "網址":         _clean_url(bkt["_url"]),
            "備註":         content[:300].strip().replace("\n", " "),
            "通報日期":     bkt["_date"],
            "縣市":         city_hit,
            "參考位置":     loc_snip,
            "工程進度":     _extract_progress(bkt["_title"] + content[:200]),
            "工程現況":     content[:200].strip().replace("\n", " "),
            "(預計)完工日": _extract_completion(content),
            "來源":         bkt["_source"],
            "來源分類":     "JSON新聞資料",
        })

    results.sort(key=lambda x: x["通報日期"], reverse=True)
    dedup_mode = "LLM" if LLM_DEDUP and LLM_CFG["api_key"] else "fallback文字"
    print(f"  → JSON 命中 {len(results)} 筆（{dedup_mode}去重後，近 {NEWS_DAYS} 天）")
    return results, dedup_track


# ─────────────────────────────────────────────────────────────
# 6. Web Search（來源 B）
# ─────────────────────────────────────────────────────────────
def _search_news(query: str, max_results: int) -> list[dict]:
    """
    Google News RSS 搜尋（免 API key，穩定可靠）。
    回傳格式與舊 DDG 相容：{title, url, body, source, date}
    """
    import urllib.request, urllib.parse, xml.etree.ElementTree as ET, html
    encoded = urllib.parse.quote(query)
    url = (f"https://news.google.com/rss/search"
           f"?q={encoded}&hl=zh-TW&gl=TW&ceid=TW:zh-TW")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read()
        root = ET.fromstring(raw)
        items = root.findall(".//item")[:max_results]
        results = []
        for item in items:
            title   = html.unescape(item.findtext("title", ""))
            link    = item.findtext("link", "")
            desc    = html.unescape(item.findtext("description", ""))
            source  = item.findtext("source", "")
            pub     = item.findtext("pubDate", "")
            # strip HTML tags from description
            desc_clean = re.sub(r'<[^>]+>', '', desc).strip()
            results.append({
                "title":  title,
                "url":    link,
                "body":   desc_clean,
                "source": source,
                "date":   pub,
            })
        return results
    except Exception as e:
        print(f"Web Search 失敗（{query[:30]}…）：{type(e).__name__}: {e}")
        return []


def generate_queries(cities: list[str], build_kws: list) -> list[str]:
    """
    為多縣市動態生成查詢清單（不限區域）。
    抽取工程關鍵字中短的動作詞組合查詢字串。
    """
    action_kws = [k for k in build_kws
                  if len(k) <= 4 and not re.search(r'[A-Za-z+]', k)][:6]
    base_kws = " ".join(action_kws[:3]) if action_kws else "工程 開工 完工"
    year = date.today().year
    queries = []
    for city in cities:
        queries.extend([
            f"{city} {base_kws} {year}",
            f"{city} 捷運 新建 啟用 {year}",
            f"{city} 橋梁 道路 施工 {year}",
            f"{city} 公共工程 建設 {year}",
        ])
    return queries


def web_search_news(loc_kws: list, build_kws: list, noise_kws: list) -> tuple[list, list]:
    """動態產生查詢、執行 Web Search 並解析結果（多縣市版）"""
    queries = generate_queries(NEWS_CITIES, build_kws)
    print(f"  Web Search 查詢 {len(queries)} 組（{', '.join(NEWS_CITIES)}）")

    buckets: dict[str, dict] = {}
    dedup_track: list[dict] = []
    track_seq = 0
    total_raw = 0

    for q in queries:
        hits = _search_news(q, WEB_SEARCH_QTY)
        total_raw += len(hits)
        for h in hits:
            title   = h.get("title", "")
            clean_title = _clean_news_title(title)
            body    = h.get("body", "")
            url     = h.get("url", "")
            source  = h.get("source", "")
            dt      = h.get("date", "")[:10]

            text = title + " " + body

            # ── 過濾 noise ──
            if any(nk in text for nk in noise_kws):
                if WEB_DEBUG:
                    print(f"    [noise] {title[:40]}")
                continue

            # ── 位置比對：標題或內文含任一縣市名/縮寫即通過 ──
            has_loc = any(lk in text for lk in loc_kws)
            if not has_loc:
                if WEB_DEBUG:
                    print(f"    [no-loc] {title[:40]}")
                continue

            # ── 工程詞比對（新聞嚴格版：短詞只算標題或後接工程詞）──
            has_build = _news_has_build(title, body, build_kws)
            if not has_build:
                if WEB_DEBUG:
                    print(f"    [no-build] {title[:40]}")
                continue

            # ── LLM 後分類 ──
            ok, reason = _llm_is_construction(clean_title, body)
            case_name = _extract_case_name(clean_title, body)

            if not ok:
                track_seq += 1
                if WEB_DEBUG:
                    print(f"    [LLM-no] {title[:40]} → {reason[:40]}")
                dedup_track.append({
                    "track_id": track_seq, "title": title, "date": dt, "content_len": len(body),
                    "llm_label": "❌ 非工程", "bucket_key": "—", "dedup_reason": reason, "result": "LLM排除",
                })
                continue

            track_seq += 1
            candidate = {
                "_body": body, "_title": title, "_url": url,
                "_source": source, "_date": dt, "_case": case_name,
                "_track_id": track_seq,
            }
            bkt_key, dedup_reason = _resolve_news_bucket(candidate, buckets, "_body")

            dedup_track.append({
                "track_id": track_seq, "title": title, "date": dt, "content_len": len(body),
                "llm_label": "✅ 是工程新聞", "bucket_key": bkt_key,
                "dedup_reason": dedup_reason, "result": "PENDING",
            })

            existing = buckets.get(bkt_key)
            if existing is None or _prefer_news_candidate(candidate, existing, "_body"):
                buckets[bkt_key] = candidate

    if WEB_DEBUG:
        dedup_mode = "LLM" if LLM_DEDUP and LLM_CFG["api_key"] else "fallback文字"
        print(f"    [debug] 總抓取 {total_raw} 筆 → 命中 {len(buckets)} 組（{dedup_mode}去重）")

    # 確定 dedup_track 各筆最終結果
    winner_ids_web = {v.get("_track_id") for v in buckets.values()}
    for entry in dedup_track:
        if entry["result"] == "PENDING":
            entry["result"] = "✅ 保留" if entry.get("track_id") in winner_ids_web else "🔀 去重"

    results = []
    for bkt in buckets.values():
        body      = bkt["_body"]
        case_name = bkt["_case"]
        # 找出文中出現的第一個縣市名
        city_hit = next((c for c in NEWS_CITIES + NEWS_CITIES_SHORT if c in body + bkt["_title"]), NEWS_CITIES[0])
        idx = body.find(city_hit)
        loc_snip = body[max(0,idx-5): idx+60].strip() if idx >= 0 else city_hit

        results.append({
            "新聞標題":     bkt["_title"],
            "案名":         case_name,
            "網址":         bkt["_url"],
            "備註":         body[:300].strip().replace("\n", " "),
            "通報日期":     bkt["_date"],
            "縣市":         city_hit if city_hit in NEWS_CITIES else NEWS_CITIES[0],
            "參考位置":     loc_snip,
            "工程進度":     _extract_progress(bkt["_title"] + body[:200]),
            "工程現況":     body[:200].strip().replace("\n", " "),
            "(預計)完工日": _extract_completion(body),
            "來源":         bkt["_source"],
            "來源分類":     "Web Search",
        })

    results.sort(key=lambda x: x["通報日期"], reverse=True)
    dedup_mode = "LLM" if LLM_DEDUP and LLM_CFG["api_key"] else "fallback文字"
    print(f"  → Web Search 命中 {len(results)} 筆（{dedup_mode}去重後）")
    return results, dedup_track


def web_search_supplement(case_name: str, location: str, city: str = "") -> dict:
    """對單一工程會案件執行 Web Search，回傳補充欄位"""
    city_str = city or " ".join(NEWS_CITIES)
    query = f"{case_name} 進度 完工 位置 {city_str}"
    hits  = _search_news(query, max_results=3)
    if not hits:
        return {}
    best = max(hits, key=lambda h: len(h.get("body", "")))
    body = best.get("body", "")
    return {
        "參考位置補充": body[:80].replace("\n", " "),
        "工程進度補充": _extract_progress(best.get("title", "") + body[:200]),
        "完工日補充":   _extract_completion(body),
        "補充來源":     best.get("source", ""),
        "補充網址":     best.get("url", ""),
        "補充摘要":     body[:200].replace("\n", " "),
    }


def parse_manual_news_urls(urls: list[str], loc_kws: list, build_kws: list, noise_kws: list) -> tuple[list, list]:
    """
    手動補充新聞網址：
    - 支援單篇拆多案件（同一篇含多路線/多工程時拆成多筆）
    - 仍走既有 LLM 後分類與去重流程
    """
    valid_urls = []
    for u in urls or []:
        u = str(u or "").strip()
        if u.startswith("http://") or u.startswith("https://"):
            valid_urls.append(u)
    valid_urls = list(dict.fromkeys(valid_urls))
    if not valid_urls:
        return [], []

    buckets: dict[str, dict] = {}
    dedup_track: list[dict] = []
    track_seq = 0

    for url in valid_urls:
        try:
            article = _extract_news_article_from_url(url)
        except Exception as e:
            dedup_track.append({
                "track_id": 0,
                "title": url,
                "date": "",
                "content_len": 0,
                "llm_label": "⚠ 讀取失敗",
                "bucket_key": "—",
                "dedup_reason": f"抓取失敗:{type(e).__name__}",
                "result": "略過",
            })
            continue

        title = article.get("title", "")
        body = article.get("body", "")
        date_str = article.get("date", "")
        source = article.get("source", "")
        text = f"{title} {body}"

        noise_hit = any(nk in text for nk in noise_kws)

        ok, reason = _llm_is_construction(title, body)
        if not ok:
            dedup_track.append({
                "track_id": 0,
                "title": title or url,
                "date": date_str,
                "content_len": len(body),
                "llm_label": "❌ 非工程",
                "bucket_key": "—",
                "dedup_reason": reason,
                "result": "LLM排除",
            })
            continue

        case_names = _extract_case_name_candidates(title, body)
        if not case_names:
            case_names = [_extract_case_name(title, body)]

        for case_name in case_names:
            track_seq += 1
            route_m = re.search(r'台[0-9]{1,3}線', case_name)
            anchor = route_m.group(0) if route_m else case_name
            idx = text.find(anchor)
            snippet = text[max(0, idx - 90): idx + 260] if idx >= 0 else text[:350]
            snippet = snippet.replace("\n", " ").strip()

            city_hit = next((c for c in NEWS_CITIES if c in snippet),
                            next((c for c in NEWS_CITIES if c in text), NEWS_CITIES[0]))
            loc_hit = next((lk for lk in loc_kws if lk in snippet),
                           next((lk for lk in loc_kws if lk in text), city_hit))

            record = {
                "新聞標題":     title,
                "案名":         case_name,
                "網址":         _clean_url(url),
                "備註":         snippet[:300],
                "通報日期":     date_str,
                "縣市":         city_hit,
                "參考位置":     loc_hit,
                "工程進度":     _extract_progress(snippet),
                "工程現況":     snippet[:200],
                "(預計)完工日": _extract_completion(snippet),
                "來源":         source,
                "來源分類":     "手動網址",
            }

            candidate = {
                "_body": " ".join([
                    record.get("備註", ""),
                    record.get("工程現況", ""),
                    record.get("參考位置", ""),
                ]),
                "_title": record.get("新聞標題", ""),
                "_url": record.get("網址", ""),
                "_source": record.get("來源", ""),
                "_date": record.get("通報日期", ""),
                "_case": record.get("案名", ""),
                "_track_id": track_seq,
                "_record": record,
            }
            bkt_key, dedup_reason = _resolve_news_bucket(candidate, buckets, "_body")
            dedup_track.append({
                "track_id": track_seq,
                "title": title,
                "date": date_str,
                "content_len": len(body),
                "llm_label": "✅ 是工程新聞",
                "bucket_key": bkt_key,
                "dedup_reason": f"{dedup_reason}; 手動網址{'命中排除詞但保留處理' if noise_hit else '直接處理'}",
                "result": "PENDING",
            })

            existing = buckets.get(bkt_key)
            if existing is None or _prefer_news_candidate(candidate, existing, "_body"):
                buckets[bkt_key] = candidate

    winner_ids = {v.get("_track_id") for v in buckets.values()}
    for entry in dedup_track:
        if entry.get("result") == "PENDING":
            entry["result"] = "✅ 保留" if entry.get("track_id") in winner_ids else "🔀 去重"

    results = [v["_record"] for v in buckets.values()]
    results.sort(key=lambda x: x.get("通報日期", ""), reverse=True)
    print(f"  → 手動網址命中 {len(results)} 筆（單篇可拆多案）")
    return results, dedup_track


# ─────────────────────────────────────────────────────────────
# 7. 合併去重
# ─────────────────────────────────────────────────────────────
def merge_news_with_track(news_items: list) -> tuple[list, list]:
    """
    整合多來源新聞並做同一輪去重（JSON / Web / 手動網址可一起進來）。
    回傳 (results, dedup_track)。
    """
    buckets: dict[str, dict] = {}
    dedup_track: list[dict] = []
    track_seq = 0

    for n in news_items or []:
        record = dict(n or {})
        if not record:
            continue
        track_seq += 1
        candidate = {
            "_body": " ".join([
                str(record.get("備註", "") or ""),
                str(record.get("工程現況", "") or ""),
                str(record.get("參考位置", "") or ""),
            ]),
            "_title": record.get("新聞標題", ""),
            "_url": record.get("網址", ""),
            "_source": record.get("來源", ""),
            "_date": record.get("通報日期", ""),
            "_case": record.get("案名", ""),
            "_track_id": track_seq,
            "_record": record,
        }
        bkt_key, dedup_reason = _resolve_news_bucket(candidate, buckets, "_body")
        dedup_track.append({
            "track_id": track_seq,
            "title": record.get("新聞標題", ""),
            "date": record.get("通報日期", ""),
            "content_len": len(candidate["_body"]),
            "llm_label": "✅ 去重候選",
            "bucket_key": bkt_key,
            "dedup_reason": dedup_reason,
            "result": "PENDING",
        })
        existing = buckets.get(bkt_key)
        if existing is None:
            buckets[bkt_key] = candidate
            continue

        src_types = set([
            existing["_record"].get("來源分類", ""),
            record.get("來源分類", ""),
        ]) - {""}
        merged_src = "+".join(sorted(src_types))
        if _prefer_news_candidate(candidate, existing, "_body"):
            candidate["_record"]["來源分類"] = merged_src
            buckets[bkt_key] = candidate
        else:
            existing["_record"]["來源分類"] = merged_src

    winner_ids = {v.get("_track_id") for v in buckets.values()}
    for entry in dedup_track:
        if entry.get("result") == "PENDING":
            entry["result"] = "✅ 保留" if entry.get("track_id") in winner_ids else "🔀 去重"

    result = sorted((v["_record"] for v in buckets.values()),
                    key=lambda x: x["通報日期"], reverse=True)
    return result, dedup_track


def merge_news(json_news: list, web_news: list) -> list:
    """相容舊介面：合併兩個來源並回傳去重後清單。"""
    result, _ = merge_news_with_track((json_news or []) + (web_news or []))
    return result


def _norm_match_text(value: object) -> str:
    """正規化案名/標題，供既有案交叉比對使用。"""
    text = str(value or "")
    text = text.replace("臺", "台")
    text = re.sub(r'https?://\S+', '', text)
    text = re.sub(r'[\s　\-_｜|【】\[\]（）()「」『』:：,，.。!！?？/／]+', '', text)
    for token in ("工程", "計畫", "案", "標", "統包", "新聞", "即時", "Yahoo", "大紀元", "大纪元"):
        text = text.replace(token, "")
    return text


def _overlap_score(a: str, b: str) -> float:
    """以字元重疊與 SequenceMatcher 混合評分，避免新聞案名和正式案名格式不同時完全失配。"""
    na, nb = _norm_match_text(a), _norm_match_text(b)
    if not na or not nb:
        return 0.0
    if min(len(na), len(nb)) < 6:
        return 0.0
    if na in nb or nb in na:
        return 1.0
    set_a, set_b = set(na), set(nb)
    overlap = len(set_a & set_b) / max(1, min(len(set_a), len(set_b)))
    seq = SequenceMatcher(None, na, nb).ratio()
    return max(overlap, seq)


def _clean_excel_value(value: object) -> object:
    """將 pandas/openpyxl 常見空值轉成空字串，避免 Excel 出現 nan。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip().lower() in {"nan", "none", "nat"}:
        return ""
    return value


# ─────────────────────────────────────────────────────────────
# 8. Excel 樣式
# ─────────────────────────────────────────────────────────────
_FILL = {
    "header":  PatternFill("solid", fgColor="1F4E79"),
    "subhdr":  PatternFill("solid", fgColor="2E75B6"),
    "green":   PatternFill("solid", fgColor="E2EFDA"),
    "yellow":  PatternFill("solid", fgColor="FFF2CC"),
    "blue":    PatternFill("solid", fgColor="DEEAF1"),
    "blue2":   PatternFill("solid", fgColor="F2F9FF"),
    "gray":    PatternFill("solid", fgColor="F2F2F2"),
    "white":   PatternFill("solid", fgColor="FFFFFF"),
}
_BORDER = Border(**{s: Side(style="thin", color="BFBFBF")
                    for s in ("left","right","top","bottom")})


def _hdr(cell, light=False):
    cell.fill = _FILL["subhdr"] if light else _FILL["header"]
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _dat(cell, fill=None):
    cell.fill = fill or _FILL["white"]
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _BORDER


def _ws_title(ws, text, cols):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    c.fill = _FILL["header"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.sheet_view.showGridLines = False


def _write_header(ws, cols, row=2):
    for ci, col in enumerate(cols, 1):
        _hdr(ws.cell(row=row, column=ci, value=col), light=True)
    ws.row_dimensions[row].height = 22


def _set_widths(ws, widths: dict):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ─────────────────────────────────────────────────────────────
# 9. 工程會清冊處理（測試 1~4）
# ─────────────────────────────────────────────────────────────
def process_pcc(nlsc_df: pd.DataFrame):
    print("▶ 載入工程會 JSON...")
    d127 = load_pcc_json(JSON_127)
    d415 = load_pcc_json(JSON_415)
    ids127 = {x["projectNo"] for x in d127}
    filtered127 = {x["projectNo"] for x in filter_target(d127)}

    # 既有 NLSC 管控案名（115EMAP案管表單）
    nlsc_names  = set(nlsc_df["基本資料_案名"].dropna().astype(str))

    # 前次篩選Y / 已納入控管案名（案件管理 Excel — 2-1 廠商初篩參考資料）
    # 「已確認納入控管或前一次(114下半年)篩選Y者，不須新增，更新對應案件狀態即可」
    prev_screened_names = load_prev_screened_names(PCC_CASE_MGR)

    target415 = filter_target(d415)
    REPORT_127 = "2026-01-27"
    REPORT_415 = "2026-04-15"

    # ── 測試 1 ───────────────────────────────────────────────
    # 三層篩檢（2-1 廠商初篩邏輯）：
    #   層1 已在 NLSC 既有管控表單 → 不須新增，更新進度即可
    #   層2 前次(114下半年)篩選Y 或 已納入控管 → 不須新增，更新狀態即可
    #   層3 0127版通過關鍵字篩選（前次已處理）→ 不須新增，更新進度即可
    #   其他 → 建議新增
    t1_new, t1_exist = [], []
    for item in target415:
        case = extract_case(item, REPORT_415)
        name = case["基本資料_案名"]
        if name in nlsc_names:
            case["篩選結果"] = "已在NLSC管控"
            case["篩選說明"] = "案名已存在 NLSC 既有管控表單，不須新增，更新進度 / 完工日即可"
            t1_exist.append(case)
        elif name in prev_screened_names:
            case["篩選結果"] = "前次篩選Y／已納入控管"
            case["篩選說明"] = "案件管理表單顯示前次(114下半年)篩選Y或已納入控管，不須新增，更新狀態即可"
            t1_exist.append(case)
        elif item["projectNo"] in filtered127:
            case["篩選結果"] = "前次版本已篩選"
            case["篩選說明"] = "20260127版已通過關鍵字篩選，本次僅需更新進度 / 完工日"
            t1_exist.append(case)
        else:
            case["篩選結果"] = "★ 建議新增納入管控"
            case["篩選說明"] = "符合關鍵字 + 金額 ≥ 500萬 + 執行單位條件，建議納入管控"
            t1_new.append(case)

    # ── 測試 2 ───────────────────────────────────────────────
    map415 = {x["projectNo"]: x for x in target415}
    t2 = []
    for item127 in filter_target(d127):
        pid = item127["projectNo"]
        item415 = map415.get(pid)
        if not item415:
            continue
        bi127, bi415 = item127.get("basicInformation",{}), item415.get("basicInformation",{})
        p127 = item127.get("progressions",[])
        p415 = item415.get("progressions",[])
        l127, l415 = (p127[-1] if p127 else {}), (p415[-1] if p415 else {})

        old_p, new_p = l127.get("actualOverallProgress",""), l415.get("actualOverallProgress","")
        old_s, new_s = l127.get("status",""), l415.get("status","")
        old_c = roc_to_ad(bi127.get("scheduledCompletionDate",""))
        new_c = roc_to_ad(bi415.get("scheduledCompletionDate",""))
        changed = (old_p != new_p) or (old_c != new_c) or (old_s != new_s)

        t2.append({
            "工程會案號":              pid,
            "基本資料_案名":           item415.get("projectName",""),
            "坐落_區":                bi415.get("district",""),
            "核對追蹤_執行單位":       bi415.get("hostEntity",""),
            "核對追蹤_預算金額【千元】": bi415.get("awardingPrice") or bi415.get("totalBudget") or 0,
            "核對追蹤_參考門牌":       bi415.get("location",""),
            "20260127_進度(%)":        old_p, "20260415_進度(%)": new_p,
            "20260127_案件狀態":       old_s, "20260415_案件狀態": new_s,
            "20260127_(預計)完工日":   old_c, "20260415_(預計)完工日": new_c,
            "有無異動": "✅ 有異動" if changed else "─ 無異動",
        })

    # ── 測試 3（Web Search 補充）───────────────────────────────
    t3 = []
    for item in target415:
        case = extract_case(item, REPORT_415)
        name = case["基本資料_案名"]
        supp = web_search_supplement(name, case["核對追蹤_參考門牌"], case["坐落_縣市"])
        t3.append({
            "工程會案號":             case["工程會案號"],
            "基本資料_案名":          name,
            "核對追蹤_參考門牌_原始": case["核對追蹤_參考門牌"],
            "核對追蹤_參考門牌_補充": supp.get("參考位置補充", "（Web Search 無結果）"),
            "工程進度_補充":          supp.get("工程進度補充", ""),
            "完工日_補充":            supp.get("完工日補充", ""),
            "補充來源":               supp.get("補充來源", ""),
            "補充網址":               supp.get("補充網址", ""),
            "補充摘要":               supp.get("補充摘要", ""),
            "是否有補充": "✅ 有" if supp else "─ 無",
        })

    # ── 測試 4（案件摘要）─────────────────────────────────────
    t4 = []
    t3_map = {r["工程會案號"]: r for r in t3}
    for item in target415:
        case = extract_case(item, REPORT_415)
        supp = t3_map.get(case["工程會案號"], {})
        if supp.get("核對追蹤_參考門牌_補充"):
            case["核對追蹤_參考門牌"] = supp["核對追蹤_參考門牌_補充"]
        t4.append({
            "工程會案號":    case["工程會案號"],
            "基本資料_案名": case["基本資料_案名"],
            "坐落":          case["坐落_區"] if "北市" in case["坐落_區"] else case["坐落_縣市"] + case["坐落_區"],
            "執行單位":      case["核對追蹤_執行單位"],
            "預算（千元）":  case["核對追蹤_預算金額【千元】"],
            "進度(%)":       case["核對追蹤_進度"],
            "(預計)完工日":  case["核對追蹤_(預計)完工日"],
            "案件摘要":      auto_summary(case),
        })

    return dict(t1_new=t1_new, t1_exist=t1_exist, t2=t2, t3=t3, t4=t4,
                target415=target415)


# ─────────────────────────────────────────────────────────────
# 10. 寫出工程會 Excel
# ─────────────────────────────────────────────────────────────
def write_pcc_excel(res):
    wb = Workbook(); wb.remove(wb.active)

    # ── 執行摘要 ─────────────────────────────────────────────
    ws0 = wb.create_sheet("執行摘要")
    cities_label = "、".join(NEWS_CITIES)
    _ws_title(ws0, f"臺灣通用電子地圖異動情資蒐整—工程會清冊處理（{cities_label}）", 2)
    rows = [
        ("資料來源", "公共工程服務雲 API（output_file_all_case_format_v2）"),
        ("篩選範圍", cities_label),
        ("篩選條件", f"金額 ≥ {MIN_PRICE_K:,} 千元 + 應納入關鍵字 + 非納入關鍵字排除"),
        ("20260415 篩選通過", f"{len(res['target415'])} 筆"),
        ("測試1 ★建議新增納入管控",   f"{len(res['t1_new'])} 筆"),
        ("測試1 前次篩選Y／已納入控管（2-1：不須新增，更新狀態即可）",
         f"{sum(1 for c in res['t1_exist'] if '前次篩選Y' in c.get('篩選結果','') or '已納入控管' in c.get('篩選結果',''))} 筆"),
        ("測試1 已在NLSC管控表單（不須新增，更新進度即可）",
         f"{sum(1 for c in res['t1_exist'] if 'NLSC管控' in c.get('篩選結果',''))} 筆"),
        ("測試1 前次版本已篩選（0127版）",
         f"{sum(1 for c in res['t1_exist'] if '前次版本' in c.get('篩選結果',''))} 筆"),
        ("測試2 進度比對",   f"{len(res['t2'])} 筆"),
        ("測試3 Web Search 有補充", f"{sum(1 for r in res['t3'] if r['是否有補充']=='✅ 有')} 筆"),
        ("測試4 案件摘要",   f"{len(res['t4'])} 筆"),
        ("執行時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for ri, (k, v) in enumerate(rows, 3):
        ck = ws0.cell(row=ri, column=1, value=k)
        cv = ws0.cell(row=ri, column=2, value=v)
        bg = _FILL["blue"] if ri % 2 == 0 else _FILL["white"]
        for c in (ck, cv):
            c.fill = bg; c.border = _BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ck.font = Font(bold=True, name="Arial", size=10)
        cv.font = Font(name="Arial", size=10)
        ws0.row_dimensions[ri].height = 18
    ws0.column_dimensions["A"].width = 34
    ws0.column_dimensions["B"].width = 60

    # ── 測試1 ────────────────────────────────────────────────
    ws1 = wb.create_sheet("測試1_新增案件建議")
    ws1.freeze_panes = "C3"
    cols1 = ["篩選結果","篩選說明","工程會案號","基本資料_案名","坐落_縣市","坐落_區",
             "核對追蹤_執行單位","核對追蹤_參考門牌","核對追蹤_預算金額【千元】",
             "核對追蹤_進度","核對追蹤_(預計)完工日","核對追蹤_案件備註",
             "基本資料_通報日期","基本資料_通報來源","基本資料_聯絡電話"]
    _ws_title(ws1, f"測試1：確認20260415版工程會資料是否有建議納入管控案件（{cities_label}）", len(cols1))
    _write_header(ws1, cols1)
    for ri, case in enumerate(res["t1_new"] + res["t1_exist"], 3):
        result = case.get("篩選結果","")
        is_new      = "新增" in result
        is_prev_y   = "前次篩選Y" in result or "已納入控管" in result
        is_prev_kw  = "前次版本" in result
        is_nlsc     = "NLSC管控" in result
        if is_new:
            fg, txt_color = _FILL["green"], "375623"
        elif is_prev_y:
            fg, txt_color = _FILL["yellow"], "7F6000"
        elif is_nlsc:
            fg, txt_color = _FILL["blue"], "1F4E79"
        else:  # is_prev_kw
            fg, txt_color = _FILL["gray"], "595959"
        for ci, col in enumerate(cols1, 1):
            c = ws1.cell(row=ri, column=ci, value=case.get(col,""))
            _dat(c, fg)
            if ci == 1:
                c.font = Font(bold=True, name="Arial", size=9, color=txt_color)
        ws1.row_dimensions[ri].height = 18
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(cols1))}2"
    _set_widths(ws1, {"A":20,"B":38,"C":26,"D":42,"E":10,"F":14,"G":24,
                      "H":36,"I":18,"J":12,"K":18,"L":45,"M":16,"N":14,"O":18})

    # ── 測試2 ────────────────────────────────────────────────
    ws2 = wb.create_sheet("測試2_進度更新比對")
    ws2.freeze_panes = "C3"
    cols2 = ["有無異動","工程會案號","基本資料_案名","坐落_區","核對追蹤_執行單位",
             "核對追蹤_預算金額【千元】","20260127_進度(%)","20260415_進度(%)",
             "20260127_案件狀態","20260415_案件狀態",
             "20260127_(預計)完工日","20260415_(預計)完工日","核對追蹤_參考門牌"]
    _ws_title(ws2, f"測試2：應用20260415版工程會資料更新既有管控案件進度 / 完工日（{cities_label}）", len(cols2))
    _write_header(ws2, cols2)
    for ri, row in enumerate(res["t2"], 3):
        changed = row.get("有無異動","").startswith("✅")
        fg = _FILL["yellow"] if changed else _FILL["white"]
        for ci, col in enumerate(cols2, 1):
            c = ws2.cell(row=ri, column=ci, value=row.get(col,""))
            _dat(c, fg)
            if ci == 1:
                c.font = Font(bold=changed, name="Arial", size=9,
                              color="375623" if changed else "595959")
        ws2.row_dimensions[ri].height = 18
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(cols2))}2"
    _set_widths(ws2, {"A":14,"B":28,"C":42,"D":14,"E":24,"F":18,
                      "G":18,"H":18,"I":16,"J":16,"K":20,"L":20,"M":36})

    # ── 測試3 ────────────────────────────────────────────────
    ws3 = wb.create_sheet("測試3_爬搜補充資訊")
    ws3.freeze_panes = "C3"
    cols3 = ["是否有補充","工程會案號","基本資料_案名",
             "核對追蹤_參考門牌_原始","核對追蹤_參考門牌_補充",
             "工程進度_補充","完工日_補充","補充來源","補充網址","補充摘要"]
    _ws_title(ws3, f"測試3：應用納入管控案件進行 Web Search 補充參考位置 / 進度（{cities_label}）", len(cols3))
    _write_header(ws3, cols3)
    for ri, row in enumerate(res["t3"], 3):
        has = row.get("是否有補充","").startswith("✅")
        fg = _FILL["green"] if has else _FILL["white"]
        for ci, col in enumerate(cols3, 1):
            c = ws3.cell(row=ri, column=ci, value=row.get(col,""))
            _dat(c, fg)
        ws3.row_dimensions[ri].height = 20
    ws3.auto_filter.ref = f"A2:{get_column_letter(len(cols3))}2"
    _set_widths(ws3, {"A":12,"B":28,"C":42,"D":30,"E":40,"F":26,"G":24,"H":20,"I":18,"J":60})

    # ── 測試4 ────────────────────────────────────────────────
    ws4 = wb.create_sheet("測試4_案件摘要")
    ws4.freeze_panes = "B3"
    cols4 = ["工程會案號","基本資料_案名","坐落","執行單位",
             "預算（千元）","進度(%)","(預計)完工日","案件摘要"]
    _ws_title(ws4, f"測試4：針對納入管控案件摘錄整理工程內容（{cities_label}）", len(cols4))
    _write_header(ws4, cols4)
    for ri, row in enumerate(res["t4"], 3):
        for ci, col in enumerate(cols4, 1):
            _dat(ws4.cell(row=ri, column=ci, value=row.get(col,"")))
        ws4.row_dimensions[ri].height = 55
    _set_widths(ws4, {"A":28,"B":42,"C":18,"D":26,"E":16,"F":12,"G":16,"H":80})

    wb.save(OUT_PCC)
    print(f"✅ 工程會處理結果已儲存：{OUT_PCC.name}")


# ─────────────────────────────────────────────────────────────
# 11. 寫出新聞爬搜 Excel（含 Sheet 4 去重過程）
# ─────────────────────────────────────────────────────────────
def _write_dedup_sheet(wb, combined_track: list):
    """
    在 wb 新增「新聞去重過程」Sheet。
    3 色：淡藍(粗體)=保留, 淡黃=去重, 淡灰=LLM排除
    """
    KEEP_FILL  = PatternFill("solid", fgColor="DEEAF1")
    DEDUP_FILL = PatternFill("solid", fgColor="FFF2CC")
    LLM_FILL   = PatternFill("solid", fgColor="F2F2F2")
    BOLD_F     = Font(bold=True, name="Arial", size=9)
    NORM_F     = Font(name="Arial", size=9)

    ws = wb.create_sheet("新聞去重過程")
    ws.sheet_view.showGridLines = False

    kept_n  = sum(1 for e in combined_track if e["result"] == "✅ 保留")
    llm_n   = sum(1 for e in combined_track if e["result"] == "LLM排除")
    dedup_n = sum(1 for e in combined_track if e["result"] == "🔀 去重")

    # ── 標題 ──
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = (f"新聞去重過程  ｜  LLM後候選：{kept_n + dedup_n} 筆  →  去重後保留：{kept_n} 筆"
               f"  ／  去重：{dedup_n} 筆  ／  LLM排除：{llm_n} 筆")
    t.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    t.fill = _FILL["header"]
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── 欄位標頭 ──
    HDR_COLS = ["桶號", "新聞標題", "日期", "內容長度", "LLM判斷", "桶鍵（去重依據）", "去重判斷理由", "結果"]
    for ci, col in enumerate(HDR_COLS, 1):
        _hdr(ws.cell(row=2, column=ci, value=col), light=True)
    ws.row_dimensions[2].height = 20

    # ── 排序：LLM通過者按桶鍵排，LLM排除者置後 ──
    passed   = [e for e in combined_track if e["result"] != "LLM排除"]
    rejected = [e for e in combined_track if e["result"] == "LLM排除"]
    passed.sort(key=lambda e: (e["bucket_key"], 0 if e["result"] == "✅ 保留" else 1))

    # 指派桶號
    bucket_nums: dict[str, int] = {}
    bkt_ctr = 0
    for entry in passed:
        bk = entry["bucket_key"]
        if bk not in bucket_nums:
            bkt_ctr += 1
            bucket_nums[bk] = bkt_ctr

    # ── 寫資料 ──
    for ri, entry in enumerate(passed + rejected, 3):
        result = entry["result"]
        if result == "✅ 保留":
            fill, font = KEEP_FILL, BOLD_F
        elif result == "🔀 去重":
            fill, font = DEDUP_FILL, NORM_F
        else:
            fill, font = LLM_FILL, NORM_F

        bkt_num   = bucket_nums.get(entry["bucket_key"], "—")
        bkt_label = f"桶{bkt_num}" if isinstance(bkt_num, int) else "—"

        vals = [bkt_label, entry["title"], entry["date"],
                entry["content_len"], entry["llm_label"],
                entry["bucket_key"], entry.get("dedup_reason", ""), result]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill; c.font = font
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = _BORDER
        ws.row_dimensions[ri].height = 26

    # ── 圖例 ──
    legend_row = 3 + len(combined_track)
    ws.merge_cells(f"A{legend_row}:H{legend_row}")
    lc = ws.cell(row=legend_row, column=1,
                 value="淡藍（粗體）=✅保留（最終輸出）  淡黃=🔀去重（同桶較短副本）  淡灰=❌LLM排除（非工程新聞）  同桶號=同一案件不同報導")
    lc.font = Font(italic=True, name="Arial", size=9, color="595959")
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.fill = PatternFill("solid", fgColor="F9F9F9")
    ws.row_dimensions[legend_row].height = 18

    _set_widths(ws, {"A": 8, "B": 52, "C": 12, "D": 10, "E": 16, "F": 22, "G": 48, "H": 12})


def write_news_excel(json_news: list, web_news: list, loc_kws: list, build_kws: list,
                     json_dedup: Optional[list] = None, web_dedup: Optional[list] = None):
    all_news = merge_news(json_news, web_news)
    cutoff   = str(date.today() - timedelta(days=NEWS_DAYS))

    wb = Workbook()

    # ── Sheet 1：整合結果 ────────────────────────────────────
    ws = wb.active
    ws.title = "新聞爬搜整合結果"
    ws.freeze_panes = "B3"
    cols = ["序號","來源分類","新聞標題","案名","網址","備註",
            "通報日期","縣市","參考位置","工程進度","工程現況","(預計)完工日","來源媒體"]
    _ws_title(ws, (f"二、新聞爬搜整合結果—{'、'.join(NEWS_CITIES)}工程建設異動"
                   f"（近{NEWS_DAYS}天）｜JSON + Web Search｜{date.today()}"), len(cols))
    _write_header(ws, cols)

    for idx, n in enumerate(all_news, 1):
        ri       = idx + 2
        date_str = n["通報日期"].replace("-","")
        src_type = n.get("來源分類","")
        is_recent = n["通報日期"] >= cutoff

        if "JSON" in src_type and "Web" not in src_type:
            fg = _FILL["green"]
        elif "Web" in src_type and "JSON" not in src_type:
            fg = _FILL["blue"] if idx % 2 == 0 else _FILL["blue2"]
        else:
            fg = _FILL["yellow"]  # 整併

        vals = [f"{date_str}-{idx:03d}", src_type,
                n["新聞標題"], n["案名"], n["網址"], n["備註"],
                n["通報日期"], n["縣市"], n["參考位置"],
                n["工程進度"], n["工程現況"], n["(預計)完工日"],
                n.get("來源","")]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            _dat(c, fg)
            if ci == 5:
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if ci == 7 and is_recent:
                c.font = Font(bold=True, name="Arial", size=9, color="375623")
        ws.row_dimensions[ri].height = 58

    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    _set_widths(ws, {"A":16,"B":16,"C":46,"D":46,"E":16,"F":65,
                     "G":13,"H":10,"I":42,"J":26,"K":55,"L":26,"M":20})

    # ── Sheet 2：JSON 原始命中 ────────────────────────────────
    ws_raw = wb.create_sheet("JSON原始命中清單")
    ws_raw.freeze_panes = "B2"
    rcols = ["通報日期","新聞標題","來源媒體","案名","工程進度","(預計)完工日","備註（節錄）","網址"]
    _write_header(ws_raw, rcols, row=1)
    for ri, n in enumerate(json_news, 2):
        vals = [n["通報日期"],n["新聞標題"],n.get("來源",""),n["案名"],
                n["工程進度"],n["(預計)完工日"],n["備註"][:120],n["網址"]]
        for ci, v in enumerate(vals, 1):
            _dat(ws_raw.cell(row=ri, column=ci, value=v),
                 _FILL["green"] if ri%2==0 else _FILL["white"])
        ws_raw.row_dimensions[ri].height = 38
    _set_widths(ws_raw, {"A":13,"B":46,"C":20,"D":40,"E":20,"F":18,"G":65,"H":18})

    # ── Sheet 3：說明 ────────────────────────────────────────
    ws_note = wb.create_sheet("資料整理說明")
    ws_note.sheet_view.showGridLines = False
    ws_note.merge_cells("A1:B1")
    h = ws_note["A1"]
    h.value = "新聞爬搜資料整理原則說明"
    h.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    h.fill = _FILL["header"]
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws_note.row_dimensions[1].height = 30

    note_rows = [
        ("爬搜範圍",   f"{'、'.join(NEWS_CITIES)}工程建設相關（捷運、橋梁、公共設施等）"),
        ("時間範圍",   f"近 {NEWS_DAYS} 天（{cutoff} ~ {date.today()}）"),
        ("位置關鍵字", f"動態衍生，共 {len(loc_kws)} 個（含區名、路名等）"),
        ("工程關鍵字", f"從關鍵字.txt 載入，共 {len(build_kws)} 個應納入詞"),
        ("來源A JSON", f"每日新聞 JSON（{NEWS_JSON_GLOB}）→ {len(json_news)} 筆命中"),
        ("來源B Web",  f"Google News RSS → {len(web_news)} 筆命中"),
        ("整合後",     f"{len(all_news)} 筆（{'LLM' if LLM_DEDUP and LLM_CFG['api_key'] else 'fallback文字'} 去重後）"),
        ("底色說明",   "🟢 淺綠 = JSON ｜ 🔵 淺藍 = Web Search ｜ 🟡 淺黃 = 整併"),
        ("序號格式",   "YYYYMMDD-XXX（通報日期+流水號）"),
        ("縣市格式",   "統一以繁體「臺」表示（臺北市、臺中市等）"),
        ("整併原則",   "同案件多筆整併為最新日期+最長內容；同新聞含多案件則拆分"),
        ("整理日期",   str(date.today())),
    ]
    for ri, (k, v) in enumerate(note_rows, 3):
        ck = ws_note.cell(row=ri, column=1, value=k)
        cv = ws_note.cell(row=ri, column=2, value=v)
        bg = _FILL["blue"] if ri%2==0 else _FILL["white"]
        for c in (ck, cv):
            c.fill = bg; c.border = _BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ck.font = Font(bold=True, name="Arial", size=10)
        cv.font = Font(name="Arial", size=10)
        ws_note.row_dimensions[ri].height = 20
    ws_note.column_dimensions["A"].width = 22
    ws_note.column_dimensions["B"].width = 85

    # ── Sheet 4：新聞去重過程 ────────────────────────────────
    combined_dedup = (json_dedup or []) + (web_dedup or [])
    if combined_dedup:
        _write_dedup_sheet(wb, combined_dedup)

    wb.save(OUT_NEWS)
    print(f"✅ 新聞爬搜結果已儲存：{OUT_NEWS.name}")
    print(f"   JSON {len(json_news)} 筆 ｜ Web {len(web_news)} 筆 ｜ 整合後 {len(all_news)} 筆"
          + (f" ｜ 去重追蹤 {len(combined_dedup)} 筆" if combined_dedup else ""))


# ─────────────────────────────────────────────────────────────
# 12. 更新 NLSC 管控表單
# ─────────────────────────────────────────────────────────────
def write_nlsc_update(pcc_res: dict, news_results: list) -> list:
    """
    依 Word 文件規格，將工程會清冊(t2有異動)與新聞爬搜結果
    寫回 NLSC 管控表單對應欄位，另存為「115EMAP案管_更新版.xlsx」。

    更新欄位（依 Word 文件欄位對應）：
      ‣ 工程會 t2（有異動）→ 核對追蹤_進度(col15) + 核對追蹤_(預計)完工日(col25)
      ‣ 新聞爬搜           → 先與既有案名/別名/縣市做交叉比對；命中者更新工程現況，
                              未命中者依新聞欄位新增為候選管控列

    底色：黃 = PCC更新  ／  綠 = 新聞補充
    原始 115EMAP 檔案不異動。
    """
    import openpyxl as _ox

    OUT_NLSC_UPD  = BASE_DIR / "115EMAP案管_更新版.xlsx"
    PCC_UPD_FILL  = PatternFill("solid", fgColor="FFF2CC")   # 黃 = PCC更新
    NEWS_UPD_FILL = PatternFill("solid", fgColor="E2EFDA")   # 綠 = 新聞補充
    UPD_FONT_PCC  = Font(bold=True, name="Arial", size=9, color="7F6000")
    UPD_FONT_NEWS = Font(bold=True, name="Arial", size=9, color="375623")

    # 欄位位置（1-based）
    ID_COL = 1
    STATUS_SUM_COL = 2
    PROC_COL = 3
    TYPE_COL = 4
    CHANGE_COL = 5
    NAME_COL = 6   # 基本資料_案名
    ALIAS_COL = 7  # 基本資料_別名
    SUMMARY_COL = 8
    ATTACH_COL = 9
    REPORT_COL = 10
    SOURCE_COL = 11
    CITY_COL = 12
    REF_E_COL = 13
    REF_N_COL = 14
    PROG_COL = 15  # 核對追蹤_進度
    STAT_COL = 17  # 核對追蹤_工程現況
    HISTORY_COL = 19
    COMP_COL = 25  # 核對追蹤_(預計)完工日

    print("  載入 NLSC 原始表單（工作表2）...")
    wb = _ox.load_workbook(NLSC_FORM)
    ws = wb["工作表2"]

    # 建立「案名 → 列號」與 fuzzy 比對索引（第1列為標頭）
    name_to_row: dict[str, int] = {}
    existing_rows: list[dict] = []
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        val = row[NAME_COL - 1].value
        try:
            max_id = max(max_id, int(row[ID_COL - 1].value or 0))
        except (TypeError, ValueError):
            pass
        if val:
            row_num = row[0].row
            name = str(val).strip()
            name_to_row[name] = row_num
            alias = str(row[ALIAS_COL - 1].value or "")
            city = str(row[CITY_COL - 1].value or "")
            status = str(row[STAT_COL - 1].value or "")
            existing_rows.append({
                "row": row_num,
                "name": name,
                "alias": alias,
                "city": city,
                "haystack": " ".join([name, alias, status]),
            })

    update_log: list[dict] = []

    def _find_news_target(news: dict) -> tuple[Optional[int], float, str]:
        """新聞與既有管控案交叉比對；回傳列號、分數、命中說明。"""
        news_name = str(news.get("案名", "") or "")
        news_title = str(news.get("新聞標題", "") or "")
        news_city = str(news.get("縣市", "") or "")
        if news_name in name_to_row:
            return name_to_row[news_name], 1.0, "案名完全相同"

        best = (None, 0.0, "")
        for row in existing_rows:
            city_bonus = 0.08 if news_city and row["city"] and news_city[:2] in row["city"] else 0
            score = max(
                _overlap_score(news_name, row["name"]),
                _overlap_score(news_name, row["alias"]),
                _overlap_score(news_title, row["name"]),
                _overlap_score(news_title, row["alias"]),
            ) + city_bonus
            if score > best[1]:
                best = (row["row"], min(score, 1.0), row["name"])

        if best[0] is not None and best[1] >= 0.92:
            return best
        return None, best[1], best[2]

    def _append_news_candidate(news: dict) -> int:
        """依新聞爬搜欄位新增一列候選案件，保留來源 URL 供人工複核。"""
        nonlocal max_id
        max_id += 1
        row_num = ws.max_row + 1
        report_date = news.get("通報日期", "")
        try:
            report_date = datetime.fromisoformat(str(report_date)).date()
        except ValueError:
            pass

        values = {
            ID_COL: max_id,
            STATUS_SUM_COL: "新聞新增候選",
            PROC_COL: "11-已確認位置,先錄案",
            TYPE_COL: "",
            CHANGE_COL: "新增",
            NAME_COL: _clean_excel_value(news.get("案名", "")),
            ALIAS_COL: _clean_excel_value(news.get("新聞標題", "")),
            SUMMARY_COL: str(_clean_excel_value(news.get("備註", "")))[:500],
            ATTACH_COL: _clean_excel_value(news.get("網址", "")),
            REPORT_COL: report_date,
            SOURCE_COL: "重大工程新聞",
            CITY_COL: _clean_excel_value(news.get("縣市", "")),
            REF_E_COL: "",
            REF_N_COL: "",
            PROG_COL: _clean_excel_value(news.get("工程進度", "")),
            STAT_COL: str(_clean_excel_value(news.get("工程現況", "")) or _clean_excel_value(news.get("備註", "")) or "")[:500],
            HISTORY_COL: f"{datetime.now().strftime('%Y-%m-%d')} 新聞爬搜新增候選；來源：{news.get('來源分類','')}；URL：{news.get('網址','')}",
            COMP_COL: _clean_excel_value(news.get("(預計)完工日", "")),
        }
        for col, value in values.items():
            c = ws.cell(row=row_num, column=col, value=value)
            c.fill = NEWS_UPD_FILL
            c.font = UPD_FONT_NEWS
        return row_num

    # ── 1. 工程會 t2（有異動的案件）→ 更新進度 + 完工日 ──────
    print(f"  比對工程會有異動案件（t2）...")
    for case in pcc_res.get("t2", []):
        if not str(case.get("有無異動", "")).startswith("✅"):
            continue
        name = str(case.get("基本資料_案名", "")).strip()
        row_num = name_to_row.get(name)
        if row_num is None:
            continue

        changes = []
        new_prog = _clean_excel_value(case.get("20260415_進度(%)", ""))
        new_comp = _clean_excel_value(case.get("20260415_(預計)完工日", ""))
        old_prog = ws.cell(row=row_num, column=PROG_COL).value
        old_comp = ws.cell(row=row_num, column=COMP_COL).value

        if new_prog != "" and str(new_prog) != str(old_prog or ""):
            c = ws.cell(row=row_num, column=PROG_COL)
            c.value = new_prog; c.fill = PCC_UPD_FILL; c.font = UPD_FONT_PCC
            changes.append(f"進度: {old_prog} → {new_prog}")
        if new_comp and str(new_comp) != str(old_comp or ""):
            c = ws.cell(row=row_num, column=COMP_COL)
            c.value = new_comp; c.fill = PCC_UPD_FILL; c.font = UPD_FONT_PCC
            changes.append(f"完工日: {old_comp} → {new_comp}")

        if changes:
            update_log.append({
                "案名": name, "列號": row_num,
                "來源": "工程會PCC(測試2)",
                "更新": " ／ ".join(changes),
            })

    # ── 2. 新聞爬搜結果 → 交叉比對既有案；未命中者新增候選列 ──────
    print(f"  比對新聞爬搜結果（{len(news_results)} 筆）...")
    for news in news_results:
        name = str(news.get("案名", "")).strip()
        if not name:
            continue

        row_num, score, matched_name = _find_news_target(news)
        if row_num is None:
            row_num = _append_news_candidate(news)
            update_log.append({
                "案名": name, "列號": row_num,
                "來源": f"新聞新增候選({news.get('來源分類', '')})",
                "更新": "新增候選列：案名 / 備註 / 通報日期 / 縣市 / 參考位置 / 進度 / 工程現況 / 完工日",
            })
            continue

        changes = []
        new_status = str(_clean_excel_value(news.get("工程現況", "")) or "")[:200].strip()
        if new_status:
            c = ws.cell(row=row_num, column=STAT_COL)
            c.value = new_status; c.fill = NEWS_UPD_FILL; c.font = UPD_FONT_NEWS
            changes.append(f"工程現況（新聞比對: {matched_name or name}，score={score:.2f}）")

        new_prog = _clean_excel_value(news.get("工程進度", ""))
        if new_prog and not ws.cell(row=row_num, column=PROG_COL).value:
            c = ws.cell(row=row_num, column=PROG_COL)
            c.value = new_prog; c.fill = NEWS_UPD_FILL; c.font = UPD_FONT_NEWS
            changes.append(f"進度補充: {new_prog}")

        new_comp = _clean_excel_value(news.get("(預計)完工日", ""))
        if new_comp and not ws.cell(row=row_num, column=COMP_COL).value:
            c = ws.cell(row=row_num, column=COMP_COL)
            c.value = new_comp; c.fill = NEWS_UPD_FILL; c.font = UPD_FONT_NEWS
            changes.append(f"完工日補充: {new_comp}")

        if changes:
            update_log.append({
                "案名": name, "列號": row_num,
                "來源": f"新聞({news.get('來源分類', '')})",
                "更新": " ／ ".join(changes),
            })

    # ── 3. 新增「更新日誌」工作表 ─────────────────────────────
    if "更新日誌" in wb.sheetnames:
        del wb["更新日誌"]
    ws_log = wb.create_sheet("更新日誌")
    ws_log.sheet_view.showGridLines = False
    log_cols = ["案名", "NLSC列號", "資料來源", "更新欄位說明"]
    ws_log.merge_cells(f"A1:{get_column_letter(len(log_cols))}1")
    th = ws_log["A1"]
    th.value = (f"NLSC 管控表單更新日誌  ──  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                f"  共更新 {len(update_log)} 筆（黃=PCC  綠=新聞）")
    th.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    th.fill = _FILL["header"]
    th.alignment = Alignment(horizontal="center", vertical="center")
    ws_log.row_dimensions[1].height = 26

    for ci, col in enumerate(log_cols, 1):
        _hdr(ws_log.cell(row=2, column=ci, value=col), light=True)
    ws_log.row_dimensions[2].height = 20

    for ri, log in enumerate(update_log, 3):
        bg = _FILL["yellow"] if "PCC" in log.get("來源", "") else _FILL["green"]
        for ci, key in enumerate(["案名", "列號", "來源", "更新"], 1):
            c = ws_log.cell(row=ri, column=ci, value=log.get(key, ""))
            c.fill = bg; c.border = _BORDER
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws_log.row_dimensions[ri].height = 18

    _set_widths(ws_log, {"A": 42, "B": 8, "C": 22, "D": 65})

    wb.save(OUT_NLSC_UPD)
    print(f"✅ NLSC 更新版已儲存：{OUT_NLSC_UPD.name}（更新 {len(update_log)} 筆）")
    for log in update_log[:5]:   # 前5筆預覽
        print(f"   列{log['列號']:4d}  {log['案名'][:24]}  [{log['來源']}]  {log['更新'][:50]}")
    if len(update_log) > 5:
        print(f"   ……（共 {len(update_log)} 筆，詳見「更新日誌」工作表）")
    return update_log


# ─────────────────────────────────────────────────────────────
# 13. 主程式
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 62)
    print(f"NLSC 臺灣通用電子地圖 異動情資蒐整 Demo v4（動態版 + NLSC更新）")
    print(f"搜尋範圍（工程會 + 新聞同步）：{'、'.join(NEWS_CITIES)}  ／  {date.today()}")
    print("=" * 62)

    # ── 載入 NLSC 管控表單（共用）────────────────────────────
    print("▶ 載入 NLSC 管控表單...")
    nlsc_df = pd.read_excel(NLSC_FORM, sheet_name="工作表2")

    # ── 動態衍生關鍵字 ────────────────────────────────────────
    build_kws = KW["include"]
    noise_kws = KW["exclude"]

    # 工程會 + 新聞共用同一組多縣市位置關鍵字
    news_loc_kws = derive_news_loc_keywords(NEWS_CITIES, NEWS_CITIES_SHORT, nlsc_df)

    print(f"  位置關鍵字：{len(news_loc_kws)} 個（前8: {news_loc_kws[:8]}）")
    print(f"  工程關鍵字：{len(build_kws)} 個 ／ 排除詞：{len(noise_kws)} 個")

    # ── 一、工程會清冊處理 ────────────────────────────────────
    print("\n【一、工程會清冊處理（測試 1~4）】")
    pcc_res = process_pcc(nlsc_df)
    write_pcc_excel(pcc_res)

    # ── 二、新聞爬搜 ─────────────────────────────────────────
    print(f"\n【二、新聞爬搜（JSON + Web Search，範圍：{'+'.join(NEWS_CITIES)}）】")
    print(f"  來源A：掃描每日新聞 JSON（近 {NEWS_DAYS} 天）...")
    json_news, json_dedup = parse_news_json(news_loc_kws, build_kws, noise_kws)

    print(f"  來源B：Google News RSS Web Search...")
    web_news, web_dedup   = web_search_news(news_loc_kws, build_kws, noise_kws)

    write_news_excel(json_news, web_news, news_loc_kws, build_kws, json_dedup, web_dedup)

    # ── 三、更新 NLSC 管控表單 ───────────────────────────────
    print("\n【三、更新 NLSC 管控表單（根據工程會 t2 異動 + 新聞爬搜補充）】")
    all_news = merge_news(json_news, web_news)
    write_nlsc_update(pcc_res, all_news)

    print("\n" + "=" * 62)
    print("Demo 執行完畢！")
    print(f"  → {OUT_PCC.name}")
    print(f"  → {OUT_NEWS.name}")
    print(f"  → 115EMAP案管_更新版.xlsx")
    print("=" * 62)
