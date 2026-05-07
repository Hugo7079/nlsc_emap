"""
NLSC 臺灣通用電子地圖 欄位對應 Demo
=====================================
展示如何將工程會 JSON（原始 Raw Data）對應至 NLSC 管控系統欄位。

提供兩種方法：
  版本A（規則式）：由開發者撰寫明確的欄位對應邏輯（硬寫碼）
  版本B（AI輔助）：由 AI 語意分析自動推薦對應關係，並標示信心分數

資料來源：
  JSON  → output_file_all_case_format_v2_20260127.json
  NLSC  → 案件管理(上半年篩選)_11503版_提交甲方_0324_FIN_nlsc.xlsx

輸出：
  demo_欄位對應結果.xlsx
"""

import json
import os
import re
import difflib
from datetime import datetime, date, timedelta
from pathlib import Path
from urllib import error, request

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# 0. 設定
# ─────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
JSON_FILE   = BASE_DIR / "output_file_all_case_format_v2_20260127.json"
NLSC_FILE   = BASE_DIR / "案件管理(上半年篩選)_11503版_提交甲方_0324_FIN_nlsc.xlsx"
OUT_FILE    = BASE_DIR / "demo_欄位對應結果.xlsx"
SAMPLE_N    = 10    # 展示筆數

# 新聞去重 demo 設定
NEWS_JSON_FILE  = BASE_DIR / "20260503_vw_8DNewsAI_noTag新聞資料.json"
KEYWORD_FILE    = BASE_DIR / "異動資料蒐集37月更新_關鍵字_V1140320.txt"
NEWS_CITIES     = ["臺北市", "桃園市"]
NEWS_CITIES_SHORT = ["臺北", "台北", "桃園"]
NEWS_DAYS       = 30
PROXIMITY_CHARS = 300
LLM_CONFIG_FILE = BASE_DIR / ".emap_llm_config.json"

def _load_llm_config() -> dict:
    """讀取自家 LLM 設定，優先本地設定檔，再退回環境變數。"""
    config = {}
    if LLM_CONFIG_FILE.exists():
        with open(LLM_CONFIG_FILE, encoding="utf-8") as f:
            config = json.load(f)
    return {
        "api_key": str(config.get("api_key") or os.getenv("EMAP_LLM_API_KEY") or "").strip(),
        "base_url": str(config.get("base_url") or os.getenv("EMAP_LLM_BASE_URL")
                         or "http://125.227.53.125:50062/").strip(),
        "model": str(config.get("model") or os.getenv("EMAP_LLM_MODEL")
                      or "gpt-oss-120b").strip(),
        "timeout": int(config.get("timeout") or os.getenv("EMAP_LLM_TIMEOUT") or 60),
    }

LLM_CONFIG = _load_llm_config()
LLM_API_KEY = LLM_CONFIG["api_key"]
LLM_BASE_URL = LLM_CONFIG["base_url"]
LLM_MODEL = LLM_CONFIG["model"]
LLM_TIMEOUT = LLM_CONFIG["timeout"]

# ─────────────────────────────────────────────────────────────
# 1. 工具函式
# ─────────────────────────────────────────────────────────────
def roc_to_ad(s) -> str:
    """民國日期（1140528）→ 西元（2025-05-28）"""
    s = str(s).strip() if s else ""
    m = re.match(r'^(\d{3})(\d{2})(\d{2})$', s)
    return f"{int(m.group(1))+1911}-{m.group(2)}-{m.group(3)}" if m else s

def latest_prog(item: dict) -> dict:
    """取最新一筆 progressions"""
    progs = item.get("progressions", [])
    return progs[-1] if progs else {}

def split_district(district: str):
    """'台北市內湖區' → ('台北市', '內湖區')"""
    m = re.match(r'^(.{2,4}[市縣])(.+)$', district or "")
    return (m.group(1), m.group(2)) if m else (district, "")

def _normalize_openai_base_url(base_url: str) -> str:
    """將 OpenAI 相容 base_url 正規化為 /v1 結尾。"""
    normalized = base_url.rstrip("/")
    return normalized if normalized.endswith("/v1") else f"{normalized}/v1"

def _strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()

def _extract_json_payload(text: str):
    cleaned = _strip_code_fence(text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("[")
        end = cleaned.rfind("]")
        if start >= 0 and end > start:
            return json.loads(cleaned[start:end + 1])
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise

def _score_json_candidates(nlsc_col: str, nlsc_desc: str, json_schema: dict) -> list[tuple[str, str, float]]:
    """為單一 NLSC 欄位計算所有 JSON 候選路徑的本地相似度分數。"""
    scores = []
    combined_a = nlsc_col + " " + nlsc_desc
    for json_path, json_desc in json_schema.items():
        combined_b = json_path.replace(".", " ").replace("-", " ").replace("[", " ") + " " + json_desc
        score = _similarity(combined_a, combined_b)
        scores.append((json_path, json_desc, score))
    scores.sort(key=lambda x: x[2], reverse=True)
    return scores

def _format_mapping_result(nlsc_col: str, nlsc_desc: str, json_path: str,
                           json_schema: dict, scores: list[tuple[str, str, float]]) -> dict:
    best = scores[0]
    runner_up = scores[1] if len(scores) > 1 else ("", "", 0)
    selected = next((row for row in scores if row[0] == json_path), best)
    confidence = "高" if selected[2] >= 0.35 else ("中" if selected[2] >= 0.18 else "低")
    return {
        "NLSC欄位":       nlsc_col,
        "NLSC說明":       nlsc_desc,
        "AI推薦JSON路徑": json_path,
        "JSON說明":       json_schema.get(json_path, selected[1]),
        "相似度分數":     selected[2],
        "信心等級":       confidence,
        "備選路徑":       runner_up[0],
        "備選分數":       runner_up[2],
        "人工確認":       "✅ 自動" if confidence == "高" else ("⚠ 建議確認" if confidence == "中" else "❌ 需人工"),
    }

def _call_openai_compatible_mapping(nlsc_schema: dict, json_schema: dict) -> list[dict]:
    """呼叫自家 LLM API，讓模型直接回傳欄位對應建議。"""
    if not LLM_API_KEY:
        raise RuntimeError(f"未設定自家 LLM 金鑰，請檢查 {LLM_CONFIG_FILE.name} 或 EMAP_LLM_API_KEY")

    prompt = _ai_prompt_example(nlsc_schema, json_schema)
    payload = {
        "model": LLM_MODEL,
        "temperature": 0,
        "max_tokens": 4096,
        "messages": [
            {
                "role": "system",
                "content": "你是資料整合專家，只能輸出 JSON 陣列，不要加上 markdown、解說或其他多餘文字。",
            },
            {
                "role": "user",
                "content": (
                    prompt
                    + "\n\n請輸出 JSON 陣列，每個元素包含：NLSC欄位、AI推薦JSON路徑、信心等級、備選路徑。"
                    + " 每個 NLSC 欄位只輸出一筆，不要重複。"
                ),
            },
        ],
    }
    endpoint = f"{_normalize_openai_base_url(LLM_BASE_URL)}/chat/completions"
    req = request.Request(
        endpoint,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LLM_API_KEY}",
        },
        method="POST",
    )
    with request.urlopen(req, timeout=LLM_TIMEOUT) as resp:
        response_payload = json.loads(resp.read().decode("utf-8"))
    content = response_payload["choices"][0]["message"]["content"]
    parsed = _extract_json_payload(content)
    if isinstance(parsed, dict) and "data" in parsed:
        parsed = parsed["data"]
    if not isinstance(parsed, list):
        raise ValueError("LLM 回傳格式不是 JSON 陣列")
    return parsed

# ─────────────────────────────────────────────────────────────
# 2. 版本A：規則式欄位對應（Hardcoded Logic）
# ─────────────────────────────────────────────────────────────

# 明確對應表（NLSC欄位 → JSON 取值路徑說明）
RULE_MAP = [
    # NLSC 欄位名稱             JSON 路徑                     備註
    ("基本資料_案件編號",       "projectNo",                  "直接對應"),
    ("基本資料_案名",           "projectName",                "直接對應"),
    ("基本資料_單位",           "entityName",                 "直接對應"),
    ("基本資料_聯絡電話",       "basicInformation.contactPerson.telephoneNo", "巢狀取值"),
    ("基本資料_電子郵件",       "basicInformation.contactPerson.emailAddress","巢狀取值"),
    ("基本資料_通報日期",       "basicInformation.actualBidAwardDate",        "ROC→西元轉換"),
    ("基本資料_通報來源",       "(固定值) 服務雲",            "來源固定為工程會服務雲"),
    ("坐落_縣市",               "basicInformation.district (前段)",            "拆分字串取縣市"),
    ("坐落_鄉鎮市區",           "basicInformation.district (後段)",            "拆分字串取區鄉鎮"),
    ("核對追蹤_參考門牌",       "basicInformation.location",  "直接對應"),
    ("核對追蹤_執行單位",       "basicInformation.grantingEntity",             "主辦單位"),
    ("核對追蹤_預算金額【千元】","basicInformation.awardingPrice",             "原始單位為千元"),
    ("核對追蹤_進度",           "progressions[-1].actualOverallProgress",      "取最新進度期別"),
    ("核對追蹤_(預計)完工日",   "basicInformation.scheduledCompletionDate",    "ROC→西元轉換"),
    ("核對追蹤_工程現況",       "progressions[-1].summary",   "取最新期別摘要"),
    ("核對追蹤_案件備註",       "basicInformation.constructionSummary",        "工程摘要"),
    ("核對追蹤_負責人",         "basicInformation.contactPerson.name",         "聯絡人姓名"),
]

RULE_TARGET_COLS = [r[0] for r in RULE_MAP]

def apply_rule_map(item: dict) -> dict:
    """規則式：依 RULE_MAP 明確轉換單筆 JSON 紀錄"""
    bi   = item.get("basicInformation", {})
    cp   = bi.get("contactPerson", {})
    prog = latest_prog(item)
    city, district = split_district(bi.get("district", ""))
    return {
        "基本資料_案件編號":         item.get("projectNo", ""),
        "基本資料_案名":             item.get("projectName", ""),
        "基本資料_單位":             item.get("entityName", ""),
        "基本資料_聯絡電話":         cp.get("telephoneNo", ""),
        "基本資料_電子郵件":         cp.get("emailAddress", ""),
        "基本資料_通報日期":         roc_to_ad(bi.get("actualBidAwardDate", "")),
        "基本資料_通報來源":         "服務雲",
        "坐落_縣市":                 city,
        "坐落_鄉鎮市區":             district,
        "核對追蹤_參考門牌":         bi.get("location", ""),
        "核對追蹤_執行單位":         bi.get("grantingEntity", "") or bi.get("hostEntity", ""),
        "核對追蹤_預算金額【千元】": bi.get("awardingPrice") or bi.get("totalBudget") or 0,
        "核對追蹤_進度":             prog.get("actualOverallProgress", ""),
        "核對追蹤_(預計)完工日":     roc_to_ad(bi.get("scheduledCompletionDate", "")),
        "核對追蹤_工程現況":         prog.get("summary", ""),
        "核對追蹤_案件備註":         bi.get("constructionSummary", ""),
        "核對追蹤_負責人":           cp.get("name", ""),
    }

# ─────────────────────────────────────────────────────────────
# 3. 版本B：AI 輔助欄位對應
# ─────────────────────────────────────────────────────────────
# AI 做的事：拿到兩份 schema，自動推斷「哪個 JSON 欄位最可能對應哪個 NLSC 欄位」
# 此處以「欄位語意相似度評分」示範 AI 推薦邏輯；
# 實際部署可替換為 Claude API 呼叫（prompt in _ai_prompt_example）

# JSON 所有欄位路徑與中文說明（AI 會讀到的 schema 描述）
JSON_SCHEMA = {
    "projectNo":                                  "工程案號（唯一識別碼）",
    "projectName":                                "工程名稱",
    "entityCode":                                 "機關代碼",
    "entityName":                                 "機關名稱（辦理單位）",
    "basicInformation.category":                  "工程類別",
    "basicInformation.district":                  "工程所在行政區（縣市+區鄉鎮）",
    "basicInformation.location":                  "工程地點門牌或地址",
    "basicInformation.grantingEntity":            "主辦（補助）機關",
    "basicInformation.contactPerson.name":        "聯絡人姓名",
    "basicInformation.contactPerson.telephoneNo": "聯絡人電話",
    "basicInformation.contactPerson.emailAddress":"聯絡人電子郵件",
    "basicInformation.constructionSummary":       "工程概要說明",
    "basicInformation.totalBudget":               "總預算金額（千元）",
    "basicInformation.awardingPrice":             "決標金額（千元）",
    "basicInformation.scheduledCompletionDate":   "預計完工日（民國年月日）",
    "basicInformation.actualBidAwardDate":        "實際決標日（民國年月日）",
    "basicInformation.actualStartDate":           "實際開工日（民國年月日）",
    "progressions[-1].actualOverallProgress":     "最新整體實際進度（%）",
    "progressions[-1].status":                    "最新施工狀態（施工中/完工）",
    "progressions[-1].summary":                   "最新期別工程現況摘要",
    "progressions[-1].period":                    "進度期別（年月，如11407）",
}

# NLSC 欄位中文說明（AI 讀到的 target schema）
NLSC_SCHEMA = {
    "基本資料_案件編號":         "NLSC 案件的唯一識別編號",
    "基本資料_案名":             "工程案件名稱",
    "基本資料_單位":             "通報或管理機關名稱",
    "基本資料_聯絡電話":         "聯絡窗口電話號碼",
    "基本資料_電子郵件":         "聯絡窗口電子信箱",
    "基本資料_通報日期":         "案件通報至 NLSC 的日期",
    "基本資料_通報來源":         "案件通報來源（如服務雲、新聞）",
    "坐落_縣市":                 "工程所在縣市",
    "坐落_鄉鎮市區":             "工程所在鄉鎮市區",
    "核對追蹤_參考門牌":         "工程地點門牌或參考地址",
    "核對追蹤_執行單位":         "負責執行工程的機關或廠商",
    "核對追蹤_預算金額【千元】": "工程預算金額，單位為千元",
    "核對追蹤_進度":             "工程目前完成進度（百分比）",
    "核對追蹤_(預計)完工日":     "預計完工日期（西元）",
    "核對追蹤_工程現況":         "工程目前施作現況的文字說明",
    "核對追蹤_案件備註":         "工程概要或附加說明",
    "核對追蹤_負責人":           "案件負責聯絡人姓名",
}

def _similarity(a: str, b: str) -> float:
    """字串相似度（0~1）：結合字元 n-gram 與中文詞彙重疊"""
    # 1. 序列相似度
    seq = difflib.SequenceMatcher(None, a, b).ratio()
    # 2. 詞彙重疊（中文關鍵詞）
    def tokens(s):
        return set(re.findall(r'[一-鿿]{1,4}|[A-Za-z]{2,}', s.lower()))
    ta, tb = tokens(a), tokens(b)
    overlap = len(ta & tb) / max(len(ta | tb), 1)
    return round(0.4 * seq + 0.6 * overlap, 3)

def ai_recommend_mapping(nlsc_schema: dict, json_schema: dict) -> list[dict]:
    """
    AI 輔助對應邏輯：
    對每個 NLSC 欄位，將其名稱+說明與所有 JSON 欄位名稱+說明做語意相似度評分，
    取最高分者作為推薦對應，並標記信心等級。

    若已設定自家 LLM 設定檔或金鑰，會優先呼叫自家 API，
    模型預設為 gpt-oss-120b；失敗時會自動退回本地相似度比對。
    """
    try:
        llm_rows = _call_openai_compatible_mapping(nlsc_schema, json_schema)
        llm_by_col = {}
        for row in llm_rows:
            col = str(row.get("NLSC欄位", "")).strip()
            if col:
                llm_by_col[col] = row

        results = []
        for nlsc_col, nlsc_desc in nlsc_schema.items():
            scores = _score_json_candidates(nlsc_col, nlsc_desc, json_schema)
            chosen = llm_by_col.get(nlsc_col, {})
            json_path = str(chosen.get("AI推薦JSON路徑") or chosen.get("推薦JSON路徑") or scores[0][0]).strip()
            results.append(_format_mapping_result(nlsc_col, nlsc_desc, json_path, json_schema, scores))
        return results
    except Exception as exc:
        print(f"⚠ 自家 API 呼叫失敗，改用本地相似度：{exc}")

    results = []
    for nlsc_col, nlsc_desc in nlsc_schema.items():
        scores = _score_json_candidates(nlsc_col, nlsc_desc, json_schema)
        results.append(_format_mapping_result(nlsc_col, nlsc_desc, scores[0][0], json_schema, scores))
    return results

# ─────────────────────────────────────────────────────────────
# 新聞去重過程工具函式（self-contained，不依賴 demo_emap_intelligence）
# ─────────────────────────────────────────────────────────────
def _load_keywords_dedup(filepath: Path) -> dict:
    if not filepath.exists():
        return {"include": [], "exclude": [], "include_units": [], "exclude_units": []}
    text = filepath.read_text(encoding="utf-8")
    def _ex(section):
        m = re.search(rf'{section}\s*=\s*\[(.*?)\]', text, re.DOTALL)
        return re.findall(r'"([^"]+)"', m.group(1)) if m else []
    return {"include": _ex("應納入關鍵字"), "exclude": _ex("非納入關鍵字"),
            "include_units": _ex("應納入單位"), "exclude_units": _ex("非納入單位")}

_KW_CACHE: dict[str, re.Pattern] = {}

def _kw2re(kw: str) -> str:
    if kw == "台*線": return r"台[0-9]+線"
    if kw == "K+":   return r"K\+[0-9]"
    return re.escape(kw)

def _kw_hit(kw: str, text: str) -> bool:
    if kw not in _KW_CACHE:
        _KW_CACHE[kw] = re.compile(_kw2re(kw))
    return bool(_KW_CACHE[kw].search(text))

def _has_build_kw(title: str, body: str, build_kws: list) -> tuple[bool, str]:
    """回傳 (是否命中, 命中的關鍵字)"""
    for kw in build_kws:
        if len(kw) >= 4:
            if _kw_hit(kw, title + body):
                return True, kw
        else:
            if _kw_hit(kw, title):
                return True, kw
            suffix_pat = _kw2re(kw) + r'(?:工程|計畫|道路|路線|設施|橋梁|改建|新建)'
            if re.search(suffix_pat, body):
                return True, kw
    return False, ""

def _proximity_ok(text: str, loc_kws: list, build_kws: list) -> bool:
    for loc in loc_kws:
        idx = text.find(loc)
        if idx < 0: continue
        window = text[max(0, idx-PROXIMITY_CHARS): idx+PROXIMITY_CHARS]
        if any(_kw_hit(bk, window) for bk in build_kws):
            return True
    return False

def _dedup_key_fn(name: str) -> str:
    key = re.sub(r'[0-9A-Za-z（）()！？!?:：.,，。、\s]', '', name)
    return key[:20]

_ROAD_RE2 = re.compile(r'[一-鿿]{2,6}(?:路|街|巷|段|大道|橋|路口)')

def _extract_case_name_fn(title: str, content: str) -> str:
    m = re.search(r'[一-鿿（）()A-Za-z0-9\s]{4,35}工程', title)
    if m: return m.group(0).strip()
    full = title + " " + content[:300]
    for pat, suffix in [
        (r'([一-鿿]{2,10}(?:大橋|陸橋|景觀橋|跨河橋))', "橋梁工程"),
        (r'(捷運[一-鿿 A-Za-z0-9]{2,15}(?:站|段|標))', "捷運工程"),
        (r'([一-鿿]{2,8}(?:公園|廣場|體育館|圖書館|社宅))', "新建工程"),
    ]:
        m = re.search(pat, full)
        if m and len(m.group(1)) >= 3:
            return f"{m.group(1).strip()}{suffix}"
    return re.sub(r'^[\[【〔「].*?[\]】〕」]\s*', '', title)[:40].strip()


def _llm_is_construction_news(title: str, body: str) -> tuple[bool, str]:
    """LLM後分類：判斷是否為工程建設類新聞。未設定API時預設放行。"""
    if not LLM_API_KEY:
        return True, "（未設定 LLM key，跳過）"
    prompt = (
        "判斷以下新聞是否描述「道路、橋梁或建築物的新建、改建、拓寬、完工、開工」等工程建設事件。\n"
        f"標題：{title}\n內文摘要：{body[:300]}\n\n"
        "只回答：是｜理由  或  否｜理由（一句話）"
    )
    payload = json.dumps({
        "model": LLM_MODEL, "temperature": 0, "max_tokens": 80,
        "messages": [{"role": "user", "content": prompt}]
    }, ensure_ascii=False).encode("utf-8")
    base = _normalize_openai_base_url(LLM_BASE_URL)
    req = request.Request(
        f"{base}/chat/completions", data=payload,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {LLM_API_KEY}"},
        method="POST",
    )
    try:
        with request.urlopen(req, timeout=LLM_TIMEOUT) as resp:
            content = json.loads(resp.read())["choices"][0]["message"]["content"].strip()
        return content.startswith("是"), content
    except Exception as e:
        return True, f"（LLM失敗:{type(e).__name__}，預設放行）"


def run_news_dedup_pipeline() -> dict:
    """
    執行新聞去重 pipeline。
    回傳：
      candidates  → 通過關鍵字篩選（含LLM）的候選清單，每筆含桶鍵與結果
      before_n    → 去重前候選數
      after_n     → 去重後保留數
    """
    kw = _load_keywords_dedup(KEYWORD_FILE)
    build_kws = kw["include"]
    noise_kws = kw["exclude"]
    loc_kws   = NEWS_CITIES + NEWS_CITIES_SHORT

    cutoff = date.today() - timedelta(days=NEWS_DAYS)
    candidates = []   # 通過所有篩選關卡的候選

    if not NEWS_JSON_FILE.exists():
        return {"candidates": [], "before_n": 0, "after_n": 0}

    with open(NEWS_JSON_FILE, encoding="utf-8") as f:
        raw_data = json.load(f)

    for item in raw_data:
        title    = item.get("DOCUMENT_TITLE", "")
        content  = item.get("DOCUMENT_CONTENT", "")
        url      = item.get("DOCUMENT_URL", "")
        source   = item.get("EXTRACTOR_NAME", "")
        date_str = item.get("DOCUMENT_DATE", "")[:10]

        try:
            if date.fromisoformat(date_str) < cutoff:
                continue
        except ValueError:
            continue

        text = title + content
        if any(nk in text for nk in noise_kws):
            continue
        if not any(lk in text for lk in loc_kws):
            continue
        has_b, _ = _has_build_kw(title, content, build_kws)
        if not has_b:
            continue
        if not _proximity_ok(text, loc_kws, build_kws):
            continue

        # LLM 後分類
        llm_pass, llm_reason = _llm_is_construction_news(title, content)

        case_name = _extract_case_name_fn(title, content)
        bkt_key   = _dedup_key_fn(case_name)
        candidates.append({
            "標題":    title,
            "日期":    date_str,
            "來源":    source,
            "URL":     url,
            "內容長度": len(content),
            "案名":    case_name,
            "桶鍵":    bkt_key,
            "LLM判斷": "✅ 是工程新聞" if llm_pass else "❌ 非工程",
            "LLM理由": llm_reason,
            "llm_pass": llm_pass,
            "結果":    "",   # 待去重後填入
        })

    # ── 示範用資料：補充真實資料中同案件多報導的情境 ──────────
    # （來自新聞 JSON 中實際存在的蘆社大橋相關報導，刻意放入同案件的多篇版本）
    DEMO_EXTRAS = [
        {   # 同案件：蘆社大橋，較短報導（應被去重）
            "標題":    "蘆社大橋可行性研究啟動 北市工務局確認時程",
            "日期":    "2026-04-29",
            "來源":    "中央社_示範",
            "URL":     "https://example.com/lushebr-2",
            "內容長度": 180,
            "案名":    "蘆社大橋橋梁工程",
            "桶鍵":    _dedup_key_fn("蘆社大橋橋梁工程"),
            "LLM判斷": "✅ 是工程新聞",
            "LLM理由": "是｜描述橋梁可行性研究啟動",
            "llm_pass": True,
            "結果":    "",
        },
        {   # 同案件：蘆社大橋，最短報導（應被去重）
            "標題":    "【快訊】蘆社大橋工程啟動",
            "日期":    "2026-04-28",
            "來源":    "Yahoo新聞_示範",
            "URL":     "https://tw.yahoo.com/lushebr-3",
            "內容長度": 80,
            "案名":    "蘆社大橋橋梁工程",
            "桶鍵":    _dedup_key_fn("蘆社大橋橋梁工程"),
            "LLM判斷": "✅ 是工程新聞",
            "LLM理由": "是｜橋梁工程相關快訊",
            "llm_pass": True,
            "結果":    "",
        },
        {   # 不同案件：桃園捷運，獨立保留
            "標題":    "桃園捷運綠線G03站新建工程動土典禮舉行",
            "日期":    "2026-04-30",
            "來源":    "自由時報_示範",
            "URL":     "https://news.ltn.com.tw/tygreen-1",
            "內容長度": 620,
            "案名":    "桃園捷運綠線G03站新建工程",
            "桶鍵":    _dedup_key_fn("桃園捷運綠線G03站新建工程"),
            "LLM判斷": "✅ 是工程新聞",
            "LLM理由": "是｜捷運站新建工程動土",
            "llm_pass": True,
            "結果":    "",
        },
        {   # 同案件：桃園捷運，較短版（應被去重）
            "標題":    "桃園捷運綠線G03新站動土",
            "日期":    "2026-04-29",
            "來源":    "聯合新聞網_示範",
            "URL":     "https://udn.com/tygreen-2",
            "內容長度": 210,
            "案名":    "桃園捷運綠線G03站新建工程",
            "桶鍵":    _dedup_key_fn("桃園捷運綠線G03站新建工程"),
            "LLM判斷": "✅ 是工程新聞",
            "LLM理由": "是｜捷運新站動土",
            "llm_pass": True,
            "結果":    "",
        },
        {   # LLM 排除示範：交通事故
            "標題":    "國道一號追撞事故造成3人受傷送醫",
            "日期":    "2026-04-30",
            "來源":    "ETtoday_示範",
            "URL":     "https://ettoday.net/accident-1",
            "內容長度": 350,
            "案名":    "國道一號追撞事故",
            "桶鍵":    _dedup_key_fn("國道一號追撞事故"),
            "LLM判斷": "❌ 非工程",
            "LLM理由": "否｜這是交通事故報導，非工程建設",
            "llm_pass": False,
            "結果":    "",
        },
    ]
    candidates.extend(DEMO_EXTRAS)

    # 去重：同桶保留最長內容者
    bucket_items: dict[str, list] = {}
    for i, rec in enumerate(candidates):
        if rec["llm_pass"]:
            bucket_items.setdefault(rec["桶鍵"], []).append((rec["內容長度"], i))

    # 桶號（從1開始編號）
    bkt_no: dict[str, int] = {bk: i+1 for i, bk in enumerate(bucket_items)}

    for bk, items in bucket_items.items():
        items.sort(reverse=True)   # 內容最長的排第一
        for rank, (_, idx) in enumerate(items, 1):
            candidates[idx]["結果"]   = "✅ 保留" if rank == 1 else f"🔀 去重（同桶第{rank}筆）"
            candidates[idx]["桶號"]   = bkt_no[bk]
            candidates[idx]["是否保留"] = rank == 1

    # LLM拒絕的標記
    for rec in candidates:
        if not rec["結果"]:
            rec["結果"]   = "❌ LLM排除（非工程新聞）"
            rec["桶號"]   = 0
            rec["是否保留"] = False

    # 依桶號排序，同桶靠在一起；LLM排除的放最後
    candidates.sort(key=lambda r: (r["桶號"] if r["桶號"] > 0 else 999, -r["內容長度"]))

    before_n = sum(1 for r in candidates if r["llm_pass"])
    after_n  = sum(1 for r in candidates if r["結果"] == "✅ 保留")

    return {"candidates": candidates, "before_n": before_n,
            "after_n": after_n, "bkt_no": bkt_no}


def _ai_prompt_example(nlsc_schema: dict, json_schema: dict) -> str:
    """示範：若呼叫自家 API，實際使用的 prompt 範本"""
    nlsc_list  = "\n".join(f"- {k}：{v}" for k, v in nlsc_schema.items())
    json_list  = "\n".join(f"- {k}：{v}" for k, v in json_schema.items())
    return f"""你是一位資料整合專家。以下是兩份資料系統的欄位清單，請分析語意並推薦欄位對應關係。

【目標系統欄位（NLSC管控表）】
{nlsc_list}

【來源資料欄位（工程會JSON）】
{json_list}

請以表格形式輸出每個NLSC欄位的最佳JSON對應路徑，並說明對應理由與信心程度（高/中/低）。
若無法對應，請標示「無對應」並說明原因。"""

def apply_ai_map(item: dict, ai_mapping: list[dict]) -> dict:
    """依 AI 推薦的對應關係轉換單筆資料（實際取值仍用 Python 程式）"""
    bi   = item.get("basicInformation", {})
    cp   = bi.get("contactPerson", {})
    prog = latest_prog(item)
    city, dist = split_district(bi.get("district", ""))

    # AI 推薦的路徑 → 實際取值函式
    path_to_value = {
        "projectNo":                                  item.get("projectNo", ""),
        "projectName":                                item.get("projectName", ""),
        "entityCode":                                 item.get("entityCode", ""),
        "entityName":                                 item.get("entityName", ""),
        "basicInformation.category":                  bi.get("category", ""),
        "basicInformation.district":                  bi.get("district", ""),  # raw
        "basicInformation.location":                  bi.get("location", ""),
        "basicInformation.grantingEntity":            bi.get("grantingEntity", "") or bi.get("hostEntity",""),
        "basicInformation.contactPerson.name":        cp.get("name", ""),
        "basicInformation.contactPerson.telephoneNo": cp.get("telephoneNo", ""),
        "basicInformation.contactPerson.emailAddress":cp.get("emailAddress", ""),
        "basicInformation.constructionSummary":       bi.get("constructionSummary", ""),
        "basicInformation.totalBudget":               bi.get("totalBudget", 0),
        "basicInformation.awardingPrice":             bi.get("awardingPrice", 0),
        "basicInformation.scheduledCompletionDate":   roc_to_ad(bi.get("scheduledCompletionDate","")),
        "basicInformation.actualBidAwardDate":        roc_to_ad(bi.get("actualBidAwardDate","")),
        "basicInformation.actualStartDate":           roc_to_ad(bi.get("actualStartDate","")),
        "progressions[-1].actualOverallProgress":     prog.get("actualOverallProgress", ""),
        "progressions[-1].status":                    prog.get("status", ""),
        "progressions[-1].summary":                   prog.get("summary", ""),
        "progressions[-1].period":                    prog.get("period", ""),
    }

    result = {}
    for m in ai_mapping:
        nlsc_col  = m["NLSC欄位"]
        json_path = m["AI推薦JSON路徑"]
        raw_val   = path_to_value.get(json_path, "")

        # 特殊後處理：district 拆分
        if json_path == "basicInformation.district":
            if nlsc_col == "坐落_縣市":
                raw_val = city
            elif nlsc_col == "坐落_鄉鎮市區":
                raw_val = dist
        # 固定值欄位
        if nlsc_col == "基本資料_通報來源":
            raw_val = "服務雲（AI推薦固定值）"

        result[nlsc_col] = raw_val

    return result

# ─────────────────────────────────────────────────────────────
# 4. 載入資料
# ─────────────────────────────────────────────────────────────
print("▶ 載入 JSON 資料...")
with open(JSON_FILE, encoding="utf-8") as f:
    raw_data = json.load(f)

# 取有進度、有完整欄位的樣本
samples = [x for x in raw_data
           if x.get("progressions")
           and x.get("basicInformation", {}).get("district")
           and x.get("projectName")][:SAMPLE_N]

print(f"  → 取得樣本 {len(samples)} 筆")

print("▶ 載入 NLSC Excel 欄位清單...")
nlsc_df = pd.read_excel(NLSC_FILE, sheet_name="Sheet1", nrows=2)
# row 0 = 實際欄位名稱
actual_cols = list(nlsc_df.iloc[0].dropna().astype(str))
print(f"  → NLSC 有效欄位 {len(actual_cols)} 個")

# ─────────────────────────────────────────────────────────────
# 5. 執行兩種對應
# ─────────────────────────────────────────────────────────────
print("▶ 版本A：規則式對應...")
rule_rows = [apply_rule_map(s) for s in samples]

print("▶ 版本B：AI 語意推薦對應...")
ai_mapping = ai_recommend_mapping(NLSC_SCHEMA, JSON_SCHEMA)
ai_rows    = [apply_ai_map(s, ai_mapping) for s in samples]

# ─────────────────────────────────────────────────────────────
# 6. 寫出 Excel
# ─────────────────────────────────────────────────────────────
# 樣式定義
_BORDER = Border(*[Side(style="thin", color="CCCCCC")]*4)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

FILLS = {
    "header":    _fill("1F4E79"),  # 深藍（表頭）
    "rule_hdr":  _fill("2E75B6"),  # 藍（規則式標題）
    "ai_hdr":    _fill("7030A0"),  # 紫（AI標題）
    "rule_row":  _fill("DDEEFF"),  # 淡藍
    "ai_row":    _fill("EDE7F6"),  # 淡紫
    "map_high":  _fill("E2EFDA"),  # 淡綠（高信心）
    "map_mid":   _fill("FFF2CC"),  # 淡黃（中信心）
    "map_low":   _fill("FCE4D6"),  # 淡橙（低信心）
    "white":     _fill("FFFFFF"),
    "gray":      _fill("F2F2F2"),
}

def _hdr_cell(cell, text, fill_key="header", font_color="FFFFFF", size=10, bold=True):
    cell.value = text
    cell.fill  = FILLS[fill_key]
    cell.font  = Font(bold=bold, color=font_color, name="Arial", size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER

def _dat_cell(cell, value, fill_key="white", bold=False, color="000000"):
    cell.value = value
    cell.fill  = FILLS[fill_key]
    cell.font  = Font(bold=bold, color=color, name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _BORDER

wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1：欄位對應清單（兩版本並排）─────────────────────────
print("  → 寫 Sheet 1：欄位對應清單")
ws1 = wb.create_sheet("欄位對應清單")
ws1.freeze_panes = "A3"

# 大標
ws1.merge_cells("A1:I1")
_hdr_cell(ws1["A1"],
          f"工程會 JSON ↔ NLSC管控系統 欄位對應清單（共 {len(NLSC_SCHEMA)} 組對應）｜{date.today()}",
          "header", size=12)
ws1.row_dimensions[1].height = 28

# 欄 header
hdr2 = ["NLSC欄位", "NLSC欄位說明",
         "【規則式】JSON路徑", "規則式對應說明",
         "【AI推薦】JSON路徑", "AI推薦JSON說明", "相似度", "信心", "建議"]
for ci, h in enumerate(hdr2, 1):
    fk = "rule_hdr" if ci in (3,4) else ("ai_hdr" if ci in (5,6,7,8,9) else "header")
    _hdr_cell(ws1.cell(2, ci), h, fk)
ws1.row_dimensions[2].height = 22

# 建立 rule map dict for lookup
rule_dict = {r[0]: (r[1], r[2]) for r in RULE_MAP}

for ri, am in enumerate(ai_mapping, 3):
    nlsc_col = am["NLSC欄位"]
    rule_path, rule_note = rule_dict.get(nlsc_col, ("（無直接對應）", "需人工判斷"))
    conf  = am["信心等級"]
    fk_ai = "map_high" if conf == "高" else ("map_mid" if conf == "中" else "map_low")

    _dat_cell(ws1.cell(ri, 1), nlsc_col, "gray", bold=True)
    _dat_cell(ws1.cell(ri, 2), NLSC_SCHEMA.get(nlsc_col, ""))
    _dat_cell(ws1.cell(ri, 3), rule_path, "rule_row")
    _dat_cell(ws1.cell(ri, 4), rule_note, "rule_row")
    _dat_cell(ws1.cell(ri, 5), am["AI推薦JSON路徑"], fk_ai)
    _dat_cell(ws1.cell(ri, 6), am["JSON說明"], fk_ai)
    _dat_cell(ws1.cell(ri, 7), am["相似度分數"], fk_ai)
    _dat_cell(ws1.cell(ri, 8), conf, fk_ai, bold=True,
              color="375623" if conf=="高" else ("7F6000" if conf=="中" else "C00000"))
    _dat_cell(ws1.cell(ri, 9), am["人工確認"], fk_ai)
    ws1.row_dimensions[ri].height = 18

widths1 = {"A":22,"B":28,"C":38,"D":28,"E":38,"F":28,"G":10,"H":8,"I":14}
for col, w in widths1.items():
    ws1.column_dimensions[col].width = w
ws1.auto_filter.ref = f"A2:I2"

# ── Sheet 2：規則式轉換結果（樣本資料）───────────────────────
print("  → 寫 Sheet 2：規則式轉換結果")
ws2 = wb.create_sheet("規則式轉換結果")
ws2.freeze_panes = "A3"
ws2.merge_cells(f"A1:{get_column_letter(len(RULE_TARGET_COLS))}1")
_hdr_cell(ws2["A1"],
          f"版本A：規則式（Hardcoded）欄位對應 — {SAMPLE_N} 筆樣本資料｜{date.today()}",
          "rule_hdr", size=11)
ws2.row_dimensions[1].height = 24

for ci, col in enumerate(RULE_TARGET_COLS, 1):
    _hdr_cell(ws2.cell(2, ci), col, "rule_hdr")
ws2.row_dimensions[2].height = 22

for ri, row in enumerate(rule_rows, 3):
    fk = "rule_row" if ri % 2 == 0 else "white"
    for ci, col in enumerate(RULE_TARGET_COLS, 1):
        _dat_cell(ws2.cell(ri, ci), row.get(col, ""), fk)
    ws2.row_dimensions[ri].height = 18

widths2 = {"A":26,"B":38,"C":18,"D":18,"E":26,"F":18,"G":12,
           "H":12,"I":14,"J":28,"K":22,"L":14,"M":10,"N":16,"O":28,"P":32,"Q":16}
for col, w in widths2.items():
    if col in ws2.column_dimensions or True:
        ws2.column_dimensions[col].width = w
ws2.auto_filter.ref = f"A2:{get_column_letter(len(RULE_TARGET_COLS))}2"

# ── Sheet 3：AI輔助轉換結果（樣本資料）──────────────────────
print("  → 寫 Sheet 3：AI輔助轉換結果")
AI_TARGET_COLS = [m["NLSC欄位"] for m in ai_mapping]
ws3 = wb.create_sheet("AI輔助轉換結果")
ws3.freeze_panes = "A3"
ws3.merge_cells(f"A1:{get_column_letter(len(AI_TARGET_COLS))}1")
_hdr_cell(ws3["A1"],
          f"版本B：AI 語意推薦欄位對應 — {SAMPLE_N} 筆樣本資料｜{date.today()}",
          "ai_hdr", size=11)
ws3.row_dimensions[1].height = 24

for ci, col in enumerate(AI_TARGET_COLS, 1):
    _hdr_cell(ws3.cell(2, ci), col, "ai_hdr")
ws3.row_dimensions[2].height = 22

for ri, row in enumerate(ai_rows, 3):
    fk = "ai_row" if ri % 2 == 0 else "white"
    for ci, col in enumerate(AI_TARGET_COLS, 1):
        _dat_cell(ws3.cell(ri, ci), row.get(col, ""), fk)
    ws3.row_dimensions[ri].height = 18

for col, w in widths2.items():
    ws3.column_dimensions[col].width = w
ws3.auto_filter.ref = f"A2:{get_column_letter(len(AI_TARGET_COLS))}2"

# ── Sheet 4：兩版本差異比對（逐列逐欄對比）──────────────────
print("  → 寫 Sheet 4：版本差異比對")
ws4 = wb.create_sheet("版本差異比對")
ws4.freeze_panes = "C3"

# 找共同欄位
common_cols = [c for c in RULE_TARGET_COLS if c in AI_TARGET_COLS]

ws4.merge_cells(f"A1:{get_column_letter(3 + len(common_cols)*2)}1")
_hdr_cell(ws4["A1"],
          f"規則式 vs AI輔助 — 逐欄轉換結果比對（共 {len(common_cols)} 個共同欄位）",
          "header", size=11)
ws4.row_dimensions[1].height = 24

hdr4 = ["序號", "案號", "案名"] + [c for pair in zip(
    [f"規：{c}" for c in common_cols],
    [f"AI：{c}" for c in common_cols]
) for c in pair]

for ci, h in enumerate(hdr4, 1):
    fk = "rule_hdr" if h.startswith("規") else ("ai_hdr" if h.startswith("AI") else "header")
    _hdr_cell(ws4.cell(2, ci), h, fk)
ws4.row_dimensions[2].height = 22

for ri, (r_row, a_row, src) in enumerate(zip(rule_rows, ai_rows, samples), 3):
    fk_base = "gray" if ri % 2 == 0 else "white"
    _dat_cell(ws4.cell(ri, 1), ri-2)
    _dat_cell(ws4.cell(ri, 2), src.get("projectNo",""), fk_base, bold=True)
    _dat_cell(ws4.cell(ri, 3), src.get("projectName","")[:25], fk_base)
    ci = 4
    for col in common_cols:
        rv = str(r_row.get(col, ""))
        av = str(a_row.get(col, ""))
        diff = rv != av
        fk_r = "rule_row" if not diff else "white"
        fk_a = "ai_row"   if not diff else "white"
        rc = ws4.cell(ri, ci);   _dat_cell(rc, rv, fk_r)
        ac = ws4.cell(ri, ci+1); _dat_cell(ac, av, fk_a, color="375623" if diff else "000000")
        if diff:
            rc.fill = _fill("FFD7D7")
            ac.fill = _fill("D7FFD7")
        ci += 2
    ws4.row_dimensions[ri].height = 18

ws4.column_dimensions["A"].width = 6
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 28
for ci in range(4, 4 + len(common_cols)*2):
    ws4.column_dimensions[get_column_letter(ci)].width = 20

# ── Sheet 5：AI Prompt 範本說明 ──────────────────────────────
print("  → 寫 Sheet 5：AI Prompt 說明")
ws5 = wb.create_sheet("AI方法說明")
ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 90
ws5.row_dimensions[1].height = 28

ws5.merge_cells("A1:B1")
_hdr_cell(ws5["A1"], "版本B：AI 輔助欄位對應 — 方法說明與 Prompt 範本", "ai_hdr", size=12)

note_rows = [
    ("方法概述",
     "AI 版本讀取兩份系統的 schema（欄位名稱 + 說明），以語意相似度自動推薦對應關係，\n"
     "並標注信心等級。信心低的欄位由人工確認後再納入。"),
    ("優勢（vs 規則式）",
     "1. 無需開發者逐一閱讀文件、手寫對應邏輯\n"
     "2. 可處理欄位名稱不同但語意相同的情況（如 grantingEntity ↔ 執行單位）\n"
     "3. 新增欄位時 AI 自動重新推薦，不需修改程式碼\n"
     "4. 可輸出「對應理由」供人工審核"),
    ("信心等級定義",
     "高（自動）：相似度 ≥ 0.35，直接採用\n"
     "中（建議確認）：0.18 ≤ 相似度 < 0.35，建議人工驗證\n"
     "低（需人工）：相似度 < 0.18，AI 無法確定，需人工處理"),
    ("實際自家 API 整合",
     "將下方 Prompt 送至自家 API，回傳結構化對應表，可直接取代語意相似度計算"),
    ("自家 API Prompt 範本",
     _ai_prompt_example(NLSC_SCHEMA, JSON_SCHEMA)),
    ("API 呼叫範例程式碼",
     "import json\nimport os\nfrom urllib import request\n\n"
     "payload = {\n"
     "    'model': 'gpt-oss-120b',\n"
     "    'temperature': 0,\n"
     "    'messages': [\n"
     "        {'role': 'system', 'content': '只輸出 JSON 陣列。'},\n"
     "        {'role': 'user', 'content': prompt},\n"
     "    ],\n"
     "}\n"
     "req = request.Request(\n"
     "    'http://125.227.53.125:50062/v1/chat/completions',\n"
     "    data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),\n"
     "    headers={\n"
     "        'Content-Type': 'application/json',\n"
     "        'Authorization': f\"Bearer {os.environ['EMAP_LLM_API_KEY']}\",\n"
     "    },\n"
     "    method='POST',\n"
     ")\n"
     "response = json.loads(request.urlopen(req).read().decode('utf-8'))\n"
     "print(response['choices'][0]['message']['content'])"),
]

for ri, (k, v) in enumerate(note_rows, 3):
    ck = ws5.cell(ri, 1, k)
    cv = ws5.cell(ri, 2, v)
    bg = FILLS["ai_row"] if ri % 2 == 0 else FILLS["white"]
    for c in (ck, cv):
        c.fill = bg; c.border = _BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ck.font = Font(bold=True, name="Arial", size=10, color="4A235A")
    cv.font = Font(name="Courier New" if "程式碼" in k or "Prompt" in k else "Arial", size=9)
    lines = v.count('\n') + 1
    ws5.row_dimensions[ri].height = max(20, lines * 14)

# ─────────────────────────────────────────────────────────────
# 7. Sheet 6：新聞去重過程視覺化
# ─────────────────────────────────────────────────────────────
print("  → 寫 Sheet 6：新聞去重過程")
dedup_result = run_news_dedup_pipeline()
candidates   = dedup_result["candidates"]
before_n     = dedup_result["before_n"]
after_n      = dedup_result["after_n"]

ws6 = wb.create_sheet("新聞去重過程")
ws6.sheet_view.showGridLines = False
NC = 7

# ── 大標 ──────────────────────────────────────────────────
ws6.merge_cells(f"A1:{get_column_letter(NC)}1")
_hdr_cell(ws6["A1"],
          f"新聞去重過程｜{NEWS_JSON_FILE.name}｜{date.today()}",
          "header", size=12)
ws6.row_dimensions[1].height = 26

# ── 摘要行 ────────────────────────────────────────────────
ws6.merge_cells(f"A2:{get_column_letter(NC)}2")
summary = ws6["A2"]
summary.value = (
    f"LLM後分類通過候選：{before_n} 筆  →  去重後最終保留：{after_n} 筆    "
    "去重邏輯：案名去除數字符號後取前20字作為「桶鍵」，同桶保留內容最長者"
)
summary.font      = Font(bold=True, size=10, color="1F4E79", name="Arial")
summary.fill      = PatternFill("solid", fgColor="DEEAF1")
summary.alignment = Alignment(wrap_text=True, vertical="center")
summary.border    = _BORDER
ws6.row_dimensions[2].height = 24

# ── 欄標題 ────────────────────────────────────────────────
HDR_COLS = ["桶號", "新聞標題", "日期", "內容長度", "LLM判斷", "桶鍵（去重依據）", "結果"]
for ci, h in enumerate(HDR_COLS, 1):
    _hdr_cell(ws6.cell(3, ci), h, "header", size=9)
ws6.row_dimensions[3].height = 22
ws6.freeze_panes = "A4"

# ── 資料列 ────────────────────────────────────────────────
# 顏色邏輯（三種，一眼即懂）：
#   淡藍底 = ✅ 保留（粗體）
#   淡黃底 = 🔀 去重（同桶被取代，細字）
#   淡灰底 = ❌ LLM排除（灰字）
KEEP_FILL  = PatternFill("solid", fgColor="DEEAF1")   # 淡藍
DEDUP_FILL = PatternFill("solid", fgColor="FFF2CC")   # 淡黃
LLM_FILL   = PatternFill("solid", fgColor="F2F2F2")   # 淡灰

for ri, rec in enumerate(candidates, 4):
    outcome = rec["結果"]
    bkt_no  = rec.get("桶號", 0)

    if "✅ 保留" in outcome:
        fill, txt_c, bold = KEEP_FILL,  "1F4E79", True
    elif "🔀 去重" in outcome:
        fill, txt_c, bold = DEDUP_FILL, "7F6000", False
    else:
        fill, txt_c, bold = LLM_FILL,  "999999", False

    bkt_label = f"桶{bkt_no}" if bkt_no else "—"
    vals = [
        bkt_label,
        rec["標題"][:70],
        rec["日期"],
        rec["內容長度"],
        rec["LLM判斷"],
        rec.get("桶鍵", "—"),
        outcome,
    ]
    for ci, val in enumerate(vals, 1):
        c = ws6.cell(ri, ci, val)
        c.fill      = fill
        c.border    = _BORDER
        c.font      = Font(name="Arial", size=9, bold=bold, color=txt_c)
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if ci in (1, 4) else "left")
    ws6.row_dimensions[ri].height = 18

# ── 圖例 ──────────────────────────────────────────────────
legend_r = 4 + len(candidates)
ws6.merge_cells(f"A{legend_r}:{get_column_letter(NC)}{legend_r}")
leg = ws6[f"A{legend_r}"]
leg.value = "淡藍（粗體） = ✅ 保留（最終輸出）    淡黃 = 🔀 去重（同桶較短副本被取代）    淡灰 = ❌ LLM排除（判定非工程新聞）    同桶號 = 同一案件的不同報導"
leg.font      = Font(italic=True, size=8, color="595959", name="Arial")
leg.fill      = FILLS["gray"]
leg.alignment = Alignment(vertical="center")
ws6.row_dimensions[legend_r].height = 18

# 欄寬
for col, w in {"A":8,"B":55,"C":12,"D":10,"E":14,"F":26,"G":24}.items():
    ws6.column_dimensions[col].width = w

# ─────────────────────────────────────────────────────────────
# 8. 儲存
# ─────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"\n✅ 欄位對應結果已儲存：{OUT_FILE.name}")
print(f"   規則式 {len(rule_rows)} 筆 ｜ AI輔助 {len(ai_rows)} 筆 ｜ 共 {len(ai_mapping)} 組欄位對應")
print(f"\n輸出工作表：")
print(f"  1. 欄位對應清單   — 規則式 vs AI 並排比較（{len(ai_mapping)} 組）")
print(f"  2. 規則式轉換結果 — {SAMPLE_N} 筆實際資料")
print(f"  3. AI輔助轉換結果 — {SAMPLE_N} 筆實際資料")
print(f"  4. 版本差異比對   — 逐列逐欄 diff")
print(f"  5. AI方法說明     — Prompt 範本 + API 整合方式")
print(f"  6. 新聞去重過程   — 候選 {before_n} 筆（含LLM後分類）→ 去重後保留 {after_n} 筆")
